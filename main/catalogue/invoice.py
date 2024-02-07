from openpyxl import load_workbook
from mysql.connector import connect
from main.catalogue.connector import CONNECTOR
import re
from main.catalogue.query import *
from django.core.files.storage import FileSystemStorage
import json
from main.catalogue.format import *
import zlib
import zipfile
from main.catalogue.helper import *

# invoice = load_workbook(filename = 'EMPIRE INVOICE.xlsx')
dest_filename = 'invoiced.xlsx'

DATA_LETTERS = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
DATA_ALIAS = {
    "buyer_code": ['buyer'],
    "invoice_number": None,
    "sale_date": None,
    "prompt_date": None,
    "receiver_address_line1": None,
    "receiver_address_line2": None,
    "receiver_address_line3": None,
    "auction_number": None,
    "lot_number": ['lot no'],
    "invoice_number_buyer": ['invoice'],
    "mark": ['mark'],
    "warehouse": None,
    "packages": ['packages'],
    "type": ['packing type', 'packaging type'],
    "grade": ['grade'],
    "net": ['net weight', 'net weight(in kg)'],
    "gross": ['gross weight', 'gross weight(in kg)'],
    "base_price": ['base price'],
    "broker_starting_price": ['broker starting price', 'broker starting price($)'],
    "price": ['sale price', 'sale price($)'],
    "status": ['status'],
    "sold_packages": ['sold packages'],
    "certification": ['certification'],
    "manufacture_date": ['manf date'],
    "producer": ['producer'],
    "broker": ['broker'],
    "number": ["si no", "sl no"]
}

DATA_INVOICE = {
    "buyer_code": 'K5',
    "invoice_number": 'M6',
    "sale_date": 'M7',
    "prompt_date": 'M8',
    "receiver_address_line1": 'B5',
    "receiver_address_line2": 'B6',
    "receiver_address_line3": 'B7',
    "auction_number": 'K11',
    "lot_number": 'A13',
    "invoice_number_buyer": 'B13',
    "mark": 'C13',
    "warehouse": 'D13',
    "packages": 'E13',
    "type": 'F13',
    "grade": 'G13',
    "net": 'H13',
    "price": 'I13',
}

DATA_INVOICE_META = {
    "buyer_code": 'K5',
    "invoice_number": 'M6',
    "sale_date": 'M7',
    "prompt_date": 'M8',
    "receiver_address_line1": 'B5',
    "receiver_address_line2": 'B6',
    "receiver_address_line3": 'B7',
    "auction_number": 'K11',
}

DATA_INVOICE_SUBTOTALS = {
    'pkgs': 'E',
    'net': 'H',
    'amount': 'J',
    'brokerage': 'K',
    'whtax': 'L',
    'payable_amount': 'M'
}

DATA_INVOICE_TAX_SUMMARY = {
    'amount': 'A',
    'brokerage': 'C',
    'gross': 'E',
    'whtax': 'G',
    'payable_amount': 'I'
}

DATA_SALE_RELATION = {
    'lot_number': 'A',
    'invoice_number_buyer': 'B',
    'mark': 'C',
    'warehouse': 'D',
    'packages': 'E',
    'type': 'F',
    'grade': 'G',
    'net': 'H',
    'price': 'I',
    '_amount': 'J',
    '_brokerage': 'K',
    '_whtax': 'L',
    '_payable_amount': 'M'
}

DATA_SALE = {
    'lot_number': 'A',
    'invoice_number_buyer': 'B',
    'mark': 'C',
    'warehouse': 'D',
    'packages': 'E',
    'type': 'F',
    'grade': 'G',
    'net': 'H',
    'price': 'I',
    '_amount': '=H*I',
    '_brokerage': '=J*0.005',
    '_whtax': '=K*0.05',
    '_payable_amount': '=SUM(J+K)-L'
}

def InvoiceNumberQuery():
    try:
        with connect(**CONNECTOR) as connection:
            current_number = '''SELECT `number` FROM `invoice_number`'''
            with connection.cursor() as cursor:
                cursor.execute(current_number)
                row = cursor.fetchone()
                return row[0]
                        
    except Error as e:
            print(e)

INVOICE_COUNTER = InvoiceNumberQuery()
def InvoiceCounter():
    global INVOICE_COUNTER
    if(INVOICE_COUNTER < 2000):
        INVOICE_COUNTER += 1
        # if len(str(INVOICE_COUNTER)) == 1:
        #     INVOICE_COUNTER = '00' + str(INVOICE_COUNTER)
        # elif len(str(INVOICE_COUNTER)) == 2:
        #     INVOICE_COUNTER = '0' + str(INVOICE_COUNTER)
    else:
        INVOICE_COUNTER = 1
        # INVOICE_COUNTER = '00' + str(1)
    return INVOICE_COUNTER

def CloseInvoiceNumber(invoice_number, Pid):
    try:
        with connect(**CONNECTOR) as connection:
            update_lot = '''UPDATE `main_auctions` SET `invoice_number` = %s WHERE `Pid` = %s'''
            with connection.cursor() as cursor:
                cursor.execute(update_lot, (invoice_number, Pid,))
                connection.commit()

    except Error as e:
        print(e)

def GenerateInvoiceNumber(auction_number, counter):
    base = 'PRME/'
    base_conform = 'PRME_'
    return {
        'number': base + str(auction_number) + '/' + Helper.number_prefix(counter),
        'file': base_conform + str(auction_number) + '_' + Helper.number_prefix(counter)
    }
    # return {
    #     'number': base + str(auction_number) + '/' + str(InvoiceCounter()),
    #     'file': base_conform + str(auction_number) + '_' + str(InvoiceCounter())
    # }

def findAlias(list, value):
    perfect = [
        'grade',
        'mark',
        "number",
    ]
    end = len(list)
    counter = 1
    for alias in list:
        if(value in perfect):
            if(re.search(alias, value) != None):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        else:
            if(value == alias):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        counter += 1
    

def getAliasRelation(value):
    endif = 0
    for alias in DATA_ALIAS:
        endif += 1
        if(DATA_ALIAS[alias] != None):
            if(findAlias(DATA_ALIAS[alias], value.lower())):
                return alias
            else:
                if endif == len(DATA_ALIAS):
                    return alias
                else:
                    continue
        else: continue

class DataInterpretor:
    def init_data(left_bound, data_layer, file):
        left_bound = int(left_bound)
        data_layer = int(data_layer)
        folder='media/documents/sales/'
        fs = FileSystemStorage(location=folder)
        file_data = fs.open(file, 'rb+')
        WORKBOOK = load_workbook(filename = file_data )
        DATA = {}
        RELATION = []
        sheet = WORKBOOK.active
        bc = 0
        inner = list()
        endif_check = list()
        for row in sheet.values:
            allow = True
            inner_counter = 0
            for value in row:
                if(inner_counter >= left_bound-1):
                    if(value != None):
                        if(bc == data_layer-1):
                            value = re.sub(r'\t', '', value)
                            RELATION.append(value)
                        inner.append(value)
                    else:
                        inner.append(None)
                        endif_check.append(value)
                inner_counter += 1
            if(len(endif_check) >= 5 and bc >= 5):
                allow = False
                break
            elif(bc >= data_layer and len(endif_check) < 5):
                if allow: DATA[bc-(data_layer)] = row[slice(left_bound-1, len(row))]
            inner = list()
            endif_check = list()
            bc += 1
        return {
            'relation': RELATION,
            'data': DATA
        }

    def generate_data(data):
        set_data = list()
        counter = 0
        for vals in data['data'].values():
            inner_counter = 0
            inner_dict = dict()
            for inner_data in vals:
                relation = getAliasRelation(data['relation'][inner_counter])
                if(relation != False):
                    inner_dict[relation] = inner_data
                else:
                    inner_dict[relation] = None
                inner_counter += 1
            set_data.append(inner_dict)
            counter += 1
        return set_data

STACK_DATA = {}
combined = list()
LOT_STATUS_RELATION = dict()
populated = list()
BUYERS = list()
BUYERS_RELATION = dict()
BUYERS_OUTLOT = list()

def StackGenerator(input_data):
    global STACK_DATA
    global BUYERS
    global BUYERS_OUTLOT
    counter = 0
    for file in input_data:
        STACK_DATA[counter] = DataInterpretor.generate_data(
            DataInterpretor.init_data(
                file['left_bound'],
                file['data_layer'],
                file['file']
            )
        )
        counter += 1
    STACK_DATA = STACK_DATA[0]
    for lot in STACK_DATA:
        LOT_STATUS_RELATION[lot['lot_number']] = lot['status']
    for data in STACK_DATA:
        exist = list()
        inner_val = dict()
        for single in data:
            if(single in DATA_INVOICE.keys()):
                exist.append(single)
        for value in DATA_INVOICE.keys():
            if(value in exist):
                inner_val[value] = data[value]
            else:
                inner_val[value] = None
        populated.append(inner_val)
    for lot in populated:
        if lot['buyer_code'] != '':
            BUYERS.append(lot['buyer_code'])
            # BUYERS += lot['buyer_code']
    BUYERS = list(set(BUYERS))
    for buyer in BUYERS:
        BUYERS_RELATION[buyer] = list()
    for buyer in BUYERS:
        for lot in populated:
            if lot['buyer_code'] == buyer and LOT_STATUS_RELATION[lot['lot_number']].lower() == 'sold':
                BUYERS_RELATION[buyer].append(lot)
                
    for lot in populated:
        if LOT_STATUS_RELATION[lot['lot_number']].lower() == 'outlot':
            number = lot['lot_number']
            for _lot in STACK_DATA:
                if(_lot['lot_number']) == number:
                    print('-- outlot --')
                    print(_lot)
                    BUYERS_OUTLOT.append(_lot)
    
# def arrangeData():
#     global BUYERS
#     for lot in populated:
#         if lot['buyer_code'] != '': BUYERS.append(lot['buyer_code'])
#     BUYERS = set(BUYERS)
#     for buyer in BUYERS:
#         BUYERS_RELATION[buyer] = list()
#     for buyer in BUYERS:
#         for lot in populated:
#             if lot['buyer_code'] == buyer and LOT_STATUS_RELATION[lot['lot_number']].lower() == 'sold':
#                 BUYERS_RELATION[buyer].append(lot)
#             else:
#                 BUYERS_OUTLOT.append(lot)
# arrangeData()

def formatAddress(address):
    def splitAddress(val):
        data = re.split("[,]+", str(val), flags=re.IGNORECASE)
        if(data[1][0] == ' '):
            data[1] = re.sub(' ', '', data[1], 1)
        return data
    if(address !=  None):
        if(re.search(',', address)):
            # P.O. Box search
            if(re.search('P.O.Box', address)):
                address = address.replace('P.O.Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address = address.replace('P.O. Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address == address
            elif(re.search('Box', address)):
                address = address.replace('Box', 'P.O. BOX')
            else:
                address = 'P.O. BOX ' + address
            # (-) Hyphen search
            if(re.search(' - ', address)):
                address == address
            elif(re.search('- ', address)):
                address = address.replace('- ', ' - ')
            elif(re.search(' -', address)):
                address = address.replace(' -', ' - ')
            elif(re.search('-', address)):
                address = address.replace('-', ' - ')
            # Format
            address = address.upper()
            return splitAddress(address)
        else:
            if(re.search('Mombasa', address)):
                address = address.replace('Mombasa', ',MOMBASA')
            elif(re.search('Nairobi', address)):
                address = address.replace('Nairobi', ',NAIROBI')
            elif(re.search('Kericho', address)):
                address = address.replace('Kericho', ',KERICHO')
            else:
                address += ',NAIROBI'
            # P.O. Box search
            if(re.search('P.O.Box', address)):
                address = address.replace('P.O.Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address = address.replace('P.O. Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address == address
            elif(re.search('Box', address)):
                address = address.replace('Box', 'P.O. BOX')
            else:
                address = 'P.O. BOX ' + address
            # (-) Hyphen search
            if(re.search(' - ', address)):
                address == address
            elif(re.search('- ', address)):
                address = address.replace('- ', ' - ')
            elif(re.search(' -', address)):
                address = address.replace(' -', ' - ')
            elif(re.search('-', address)):
                address = address.replace('-', ' - ')
            # Format
            address = address.upper()
            return splitAddress(address)
    else:
        return ['', '']
    
def populate_number(val, level):
    data_length = len(val)
    functions = [
        'SUM', 'PRODUCT', 'DIFFERENCE', 'AVG',
    ]
    hasfn = False
    for fn in functions:
        if(re.search(fn, val)):
            hasfn = True
    counter = 0
    brack = False
    for v in val:
        if(v == '('):
            brack = True
        if v in DATA_LETTERS:
            if(not hasfn):
                val = val.replace(v, v+str(level))
            else:
                if(brack == True):
                    val = val.replace(v, v+str(level))
        if(counter == data_length-1):
            return val
        counter += 1
        
def replaceResale(v):
    resalestr = ["Resale", "resale", "RESALE"]
    for resale in resalestr:
        if resale in str(v):
            head, sep, tail = str(v).partition(resale)
            head += sep
            return str(head).replace(" ", "").replace("-Resale", "").replace("-resale", "").replace("-RESALE", "").replace("Resale", "").replace("resale", "").replace("RESALE", "")
        else:
            return v

NUMBER_FORMAT_CELLS = list()
NUMBER_VOID_CELLS = list()
COLLECTIVE_CELL_DATA = list()
def PopulateRow(sheet, level, row_data, catalogue_data):
    global NUMBER_FORMAT_CELLS
    global COLLECTIVE_CELL_DATA
    NUMBER_FORMAT_CELLS = list()
    NUMBER_VOID_CELLS = list()
    COLLECTIVE_CELL_DATA = list()
    for data in DATA_SALE:
        mark = row_data['mark']
        warehouse = DatabaseQueryProducerCompany(mark)
        if(data[0] != '_'):
            if(data == 'warehouse'):
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = GetInvoiceWarehouse(catalogue_data, replaceResale(row_data['invoice_number_buyer']), mark)
            elif(data == 'invoice_number_buyer'):
                # print(row_data['invoice_number_buyer'])
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = replaceResale(row_data['invoice_number_buyer'])
            elif(data == 'packages' or data == 'net'):
                if data == 'net':
                    if data != None:
                        print(mark)
                        print(row_data['invoice_number_buyer'])
                        print(replaceResale(row_data['invoice_number_buyer']))
                        sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = int(GetInvoiceNet(catalogue_data, replaceResale(row_data['invoice_number_buyer']), mark))
                    else:
                        sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = GetInvoiceNet(catalogue_data, replaceResale(row_data['invoice_number_buyer']), mark)
                else:
                    if data != None:
                        sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = int(row_data[data])
                    else:
                        sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = row_data[data]
            elif(data == 'type'):
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = GetInvoiceType(catalogue_data, replaceResale(row_data['invoice_number_buyer']), mark)
            else:
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = row_data[data]
            Format.formatArial11(sheet[str(str(DATA_SALE_RELATION[data])+str(level))])
            cell = populate_number(DATA_SALE_RELATION[data], level)
            if data == 'price':
                NUMBER_VOID_CELLS.append(cell)
            COLLECTIVE_CELL_DATA.append(cell)
        else:
            sheet[populate_number(DATA_SALE_RELATION[data], level)] = populate_number(DATA_SALE[data], level)
            sheet[populate_number(DATA_SALE_RELATION[data], level)].number_format = '$#,##0.00'
            Format.formatArial11(sheet[populate_number(DATA_SALE_RELATION[data], level)])
            cell = populate_number(DATA_SALE_RELATION[data], level)
            NUMBER_FORMAT_CELLS.append(cell)
            COLLECTIVE_CELL_DATA.append(cell)
    return { 
        'number_format_cells': NUMBER_FORMAT_CELLS,
        'collective_cell_data': COLLECTIVE_CELL_DATA,
        'number_void_cells': NUMBER_VOID_CELLS
    }


def GetInvoiceWarehouse(catalogue_data, invoice_number, mark):
    lot_warehouse = None
    for data in catalogue_data:
        for lot in data:
            if(lot == 'invoice_number'):
                if(str(data["invoice_number"]) == str(f'{invoice_number}||{mark}')):
                    lot_warehouse = data['warehouse']
    return lot_warehouse

def GetInvoiceNet(catalogue_data, invoice_number, mark):
    lot_warehouse = None
    for data in catalogue_data:
        for lot in data:
            if(lot == 'invoice_number'):
                if(str(data["invoice_number"]) == str(f'{invoice_number}||{mark}')):
                    lot_warehouse = data['net']
    return lot_warehouse

def GetInvoiceType(catalogue_data, invoice_number, mark):
    lot_warehouse = None
    for data in catalogue_data:
        for lot in data:
            if(lot == 'invoice_number'):
                if(str(data["invoice_number"]) == str(f'{invoice_number}||{mark}')):
                    lot_warehouse = data['type']
    return lot_warehouse

def GetInvoiceCompany(catalogue_data, invoice_number, mark):
    lot_company = None
    for data in catalogue_data:
        for lot in data:
            if(lot == 'invoice_number'):
                if(str(data["invoice_number"]) == str(f'{invoice_number}||{mark}')):
                    lot_company = data['company']
    return lot_company

def GenerateInvoice(data, custom_values, counter, buyer):
    folder='media/resources/'
    
    catalogue_data = custom_values['catalogue_data']
    folder='media/documents/catalogue_data'
    fsc = FileSystemStorage(location=folder)
    
    with fsc.open(catalogue_data, 'rb+') as fcc_file:
        file_datac = json.load(fcc_file)

    LOT_PRME = {buyer: list()}
    LOT_KTDA = {buyer: list()}

    for buyer in data:
        lot = data[buyer]
        for lot_data in lot:
            invoice_number = lot_data['invoice_number_buyer']
            mark = lot_data['mark']
            print(mark)
            lot_company = GetInvoiceCompany(file_datac, invoice_number, mark)
            if lot_company == 'KTDA':
                LOT_KTDA[buyer].append(lot_data)
            else:
                LOT_PRME[buyer].append(lot_data)
    
    def _Generate(data, buyer, custom_values, type="default"):
        folder='media/resources/'
        fs = FileSystemStorage(location=folder)
        template_default = fs.open('EMPIRE INVOICE TEMPLATE.xlsx', 'rb+')
        template_alt = fs.open('EMPIRE INVOICE TEMPLATE ALT.xlsx', 'rb+')
        
        catalogue_data = custom_values['catalogue_data']
        invoice_data = custom_values['invoice_data']
        folder='media/documents/catalogue_data'
        fsc = FileSystemStorage(location=folder)
        
        with fsc.open(catalogue_data, 'rb+') as fcc_file:
            file_datac = json.load(fcc_file)

        invoice_file = GenerateInvoiceNumber(custom_values['auction_number'], counter)['file']
        if len(LOT_PRME[buyer]) >= 1:
            invoice_file_alt = GenerateInvoiceNumber(custom_values['auction_number'], counter+1)['file']
        else:
            invoice_file_alt = GenerateInvoiceNumber(custom_values['auction_number'], counter)['file']
            
        fs_save_folder = 'media/documents/invoices/'
        
        invoice_number = GenerateInvoiceNumber(custom_values['auction_number'], counter)['number']
        if len(LOT_PRME[buyer]) >= 1:
            invoice_number_alt = GenerateInvoiceNumber(custom_values['auction_number'], counter+1)['number']
        else:
            invoice_number_alt = GenerateInvoiceNumber(custom_values['auction_number'], counter)['number']
        
        if type != 'default':
            invoice_number = invoice_number_alt
        
        sale_date = custom_values['sale_date']
        prompt_date = custom_values['prompt_date']
        if(type == 'default'):
            file_data = template_default
            focus_file = invoice_file
        else:
            file_data = template_alt
            focus_file = invoice_file_alt
        
        empire_template = load_workbook(filename = file_data)
        
        if(type == 'default'):
            _filename = 'Invoice_' + '(' + buyer + ')' + invoice_file + '.xlsx'
        else:
            _filename = 'Invoice__' + '(' + buyer + ')' + invoice_file_alt + '.xlsx'
        
        dest_save_path = fs_save_folder + _filename
        for buyer in data:
            if(buyer == 'CKLB'):
                buyer_company = DatabaseQueryBuyerCompany('CKL')
            elif(buyer == 'JFLB'):
                buyer_company = DatabaseQueryBuyerCompany('JFL')
            else: 
                buyer_company = DatabaseQueryBuyerCompany(buyer)
            if(buyer == 'CKLB'):
                buyer_address = DatabaseQueryBuyerAddress('CKL')
            elif(buyer == 'JFLB'):
                buyer_address = DatabaseQueryBuyerAddress('JFL')
            else: 
                buyer_address = DatabaseQueryBuyerAddress(buyer)
            code = buyer
            address_line1 = str(buyer_company).upper()
            address_line2 = formatAddress(buyer_address)[0]
            address_line3 = formatAddress(buyer_address)[1]
            meta_relation = {
                'buyer_code': 'BUYER CODE: ' + code,
                'invoice_number': 'INVOICE NO: ' + invoice_number,
                'sale_date': 'Sale Date: ' + sale_date,
                'prompt_date': 'Prompt Date: ' + prompt_date,
                'receiver_address_line1': address_line1,
                'receiver_address_line2': address_line2,
                'receiver_address_line3': address_line3,
                'auction_number': custom_values['auction_number_full']
            }
            lot_start = 13
            lot_limit_start = lot_start
            subtotals = 14
            tax_summary = 19
            lot = data[buyer]
            data_length = len(lot)
            subtotals += data_length-1
            tax_summary += data_length-1
            if(data_length > 1):
                empire_template.active.insert_rows(25, data_length-1)
                empire_template.active.insert_rows(lot_start, data_length-1)
            NUMBER_CELLS = list()
            COLLECTIVE_CELLS = list()
            NUMBER_VOID = list()
            for lot_data in lot:
                # NUMBER_CELLS.append(PopulateRow(empire_template.active, lot_start, lot_data, file_datac))
                CELL_DATA = PopulateRow(empire_template.active, lot_start, lot_data, file_datac)
                NUMBER_CELLS += [*CELL_DATA['number_format_cells'], *NUMBER_CELLS]
                COLLECTIVE_CELLS += [*CELL_DATA['collective_cell_data'], *COLLECTIVE_CELLS]
                NUMBER_VOID += [*CELL_DATA['number_void_cells'], *NUMBER_VOID]
                lot_start += 1
            # ------ GLOBAL FORMATTING --------
            for cell in COLLECTIVE_CELLS:
                Format.formatArial11(empire_template.active[cell])
            for cell in NUMBER_CELLS:
                empire_template.active[cell].number_format = '$#,##0.00'
            for cell in NUMBER_VOID:
                empire_template.active[cell].number_format = '#,##0.00'
            lot_end = lot_start-1
            SUMMARY_RELATION = {}
            # print(lot_end+1)
            for subtotal in DATA_INVOICE_SUBTOTALS:
                empire_template.active[DATA_INVOICE_SUBTOTALS[subtotal]+str(subtotals)] = '=SUM(' + DATA_INVOICE_SUBTOTALS[subtotal] + str(lot_limit_start) + ':' + DATA_INVOICE_SUBTOTALS[subtotal] + str(lot_end) + ')'
                SUMMARY_RELATION[subtotal] = '=' + DATA_INVOICE_SUBTOTALS[subtotal]+str(subtotals)
                Format.formatArial11Bold(empire_template.active[DATA_INVOICE_SUBTOTALS[subtotal]+str(subtotals)])
                empire_template.active[DATA_INVOICE_SUBTOTALS[subtotal]+str(subtotals)].border = medium_border
                if subtotal != 'pkgs' and subtotal != 'net':
                    empire_template.active[DATA_INVOICE_SUBTOTALS[subtotal]+str(subtotals)].number_format = '$#,##0.00'
            # print(subtotals)
            # print(tax_summary)
            # print(SUMMARY_RELATION)
            for summary in DATA_INVOICE_TAX_SUMMARY:
                if(summary != 'gross'):
                    empire_template.active[DATA_INVOICE_TAX_SUMMARY[summary]+str(tax_summary)] = SUMMARY_RELATION[summary]
                    empire_template.active[DATA_INVOICE_TAX_SUMMARY[summary]+str(tax_summary)].number_format = '$#,##0.00'
                else:
                    empire_template.active[DATA_INVOICE_TAX_SUMMARY[summary]+str(tax_summary)] = '=SUM(' + DATA_INVOICE_TAX_SUMMARY['amount'] + str(tax_summary) + ':' + DATA_INVOICE_TAX_SUMMARY['brokerage'] + str(tax_summary) + ')'
                    empire_template.active[DATA_INVOICE_TAX_SUMMARY[summary]+str(tax_summary)].number_format = '$#,##0.00'
            for meta in DATA_INVOICE_META:
                empire_template.active[DATA_INVOICE_META[meta]] = meta_relation[meta]
                
            # workbook = xlsxwriter.Workbook(dest_save_path)
            # worksheet = workbook.add_worksheet()
            # row = 1
            # col = 3
            # options = {
            #     'fill': {'none': True},
            # }
            # text = 'PRIME TEA BROKERS LIMITED\n1st Floor Suite 4, Rex House, Moi Avenue\nP.O. Box 83070 - 80100\nMOMBASA, KENYA\nTel No. +254 114 591 868\ninfo@primeteabrokers.com'
            # worksheet.insert_textbox(row, col, text, options)
                
            empire_template.active.merge_cells('A' + str(tax_summary) + ':B' + str(tax_summary))
            empire_template.active.merge_cells('C' + str(tax_summary) + ':D' + str(tax_summary))
            empire_template.active.merge_cells('E' + str(tax_summary) + ':F' + str(tax_summary))
            empire_template.active.merge_cells('G' + str(tax_summary) + ':H' + str(tax_summary))
            empire_template.active.merge_cells('I' + str(tax_summary) + ':J' + str(tax_summary))
            empire_template.active.merge_cells('A' + str(tax_summary-1) + ':B' + str(tax_summary-1))
            empire_template.active.merge_cells('C' + str(tax_summary-1) + ':D' + str(tax_summary-1))
            empire_template.active.merge_cells('E' + str(tax_summary-1) + ':F' + str(tax_summary-1))
            empire_template.active.merge_cells('G' + str(tax_summary-1) + ':H' + str(tax_summary-1))
            empire_template.active.merge_cells('I' + str(tax_summary-1) + ':J' + str(tax_summary-1))
            
            Format.formatArial11Bold(empire_template.active[('A' + str(tax_summary))])
            Format.formatArial11Bold(empire_template.active[('C' + str(tax_summary))])
            Format.formatArial11Bold(empire_template.active[('E' + str(tax_summary))])
            Format.formatArial11Bold(empire_template.active[('G' + str(tax_summary))])
            Format.formatArial11Bold(empire_template.active[('I' + str(tax_summary))])
            Format.formatArial11Bold(empire_template.active[('A' + str(tax_summary-1))])
            Format.formatArial11Bold(empire_template.active[('C' + str(tax_summary-1))])
            Format.formatArial11Bold(empire_template.active[('E' + str(tax_summary-1))])
            Format.formatArial11Bold(empire_template.active[('G' + str(tax_summary-1))])
            Format.formatArial11Bold(empire_template.active[('I' + str(tax_summary-1))])
            
            empire_template.active[('A' + str(tax_summary))].border = subtotals_border
            empire_template.active[('C' + str(tax_summary))].border = subtotals_border
            empire_template.active[('E' + str(tax_summary))].border = subtotals_border
            empire_template.active[('G' + str(tax_summary))].border = subtotals_border
            empire_template.active[('I' + str(tax_summary))].border = subtotals_border
            empire_template.active[('A' + str(tax_summary-1))].border = subtotals_border
            empire_template.active[('C' + str(tax_summary-1))].border = subtotals_border
            empire_template.active[('E' + str(tax_summary-1))].border = subtotals_border
            empire_template.active[('G' + str(tax_summary-1))].border = subtotals_border
            empire_template.active[('I' + str(tax_summary-1))].border = subtotals_border
            
            empire_template.active[('A' + str(tax_summary))].number_format = '$#,##0.00'
            empire_template.active[('C' + str(tax_summary))].number_format = '$#,##0.00'
            empire_template.active[('E' + str(tax_summary))].number_format = '$#,##0.00'
            empire_template.active[('G' + str(tax_summary))].number_format = '$#,##0.00'
            empire_template.active[('I' + str(tax_summary))].number_format = '$#,##0.00'
            empire_template.active[('A' + str(tax_summary-1))].number_format = '$#,##0.00'
            empire_template.active[('C' + str(tax_summary-1))].number_format = '$#,##0.00'
            empire_template.active[('E' + str(tax_summary-1))].number_format = '$#,##0.00'
            empire_template.active[('G' + str(tax_summary-1))].number_format = '$#,##0.00'
            empire_template.active[('I' + str(tax_summary-1))].number_format = '$#,##0.00'
            
            empire_template.active.title = focus_file
            
            empire_template.save(filename=dest_save_path)
            
            return _filename
    
    _dir = None
    _dir_alt = None
    if len(LOT_KTDA[buyer]) >= 1:
        _dir_alt = _Generate(LOT_KTDA, buyer, custom_values, "alt")
    if len(LOT_PRME[buyer]) >= 1:
        _dir = _Generate(LOT_PRME, buyer, custom_values, "default")
    
    if len(LOT_PRME[buyer]) >= 1 and len(LOT_KTDA[buyer]) >= 1:
        counter = counter+1
    else:
        counter == counter

    return {
        'file': _dir,
        'file_alt': _dir_alt,
        'counter': counter,
    }

class PopulateInvoice():
    def fill_lots(custom_values):
        counter = int(custom_values['invoice_number'])
        dirs = list()
        dirs_alt = list()
        for buyer in BUYERS_RELATION:
            buyer_dirs = GenerateInvoice({
                buyer: BUYERS_RELATION[buyer]
            }, custom_values, counter, buyer)
            print(buyer_dirs)
            if buyer_dirs['file']:
                dirs.append(buyer_dirs['file'])
            if buyer_dirs['file_alt']:
                dirs_alt.append(buyer_dirs['file_alt'])
            counter = buyer_dirs['counter']
            counter += 1
        CloseInvoiceNumber(counter, custom_values['auction_Pid'])
        print(BUYERS_OUTLOT)
        return {
            'dirs': dirs,
            'dirs_alt': dirs_alt,
            'outlot': BUYERS_OUTLOT,
        }
        
def INVOICEGENERATOR(input_data, custom_data):
    StackGenerator(input_data)
    return PopulateInvoice.fill_lots(custom_data)
