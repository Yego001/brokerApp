from numpy import number
from openpyxl import load_workbook
import pandas as pd
from mysql.connector import connect, Error
from connector import CONNECTOR
import json
import re
from threading import Timer

# data = pd.read_excel('DATA SETUP.xlsx', sheet_name=None, names=None, skiprows=2)

# parsed = pd.ExcelFile.parse('DATA SETUP.xlsx', sheet_name="BROKERS");

wb = load_workbook(filename = 'DATA SETUP.xlsx')
wbp = load_workbook(filename = 'Producer and Marks.xlsx')

sheet = wb.active

def readExcel():
    return

data_brokers = wb['BROKERS']
data_warehouses = wb['WAREHOUSES']
data_buyers = wb['BUYERS']
data_producers = wbp['PRODUCER AND MARKS']
data_producers_meta = wb['PRODUCERS']

brokers = dict()
warehouses = dict()
buyers = dict()
producers = dict()
producers_meta = dict()
producers_relation = list()

LANDLINES = {
    '020': 'Nairobi',
    '041': 'Mombasa',
    '051': 'Nakuru',
    '053': 'Eldoret',
    '057': 'Kisumu',
    '052': '_', #Kenya
    '006': '_', #Uganda
    '007': '_', #Uganda
}

TELEPHONES_VOID = list()
TELEPHONES = list()
for n in range(1, 100):
    if(len(str(n)) == 1):
        val = '0'+str(n)
    else:
        val = str(n)
    number_void = '7'+val
    number = '07'+val
    TELEPHONES.append(str(number))
    TELEPHONES_VOID.append(str(number_void))
    
def splitNumbers(value):
    return re.split("[a-z/']+", str(value)
        .replace(" ", "")
        .replace("-", "")
        .replace("C/o", "")
        .replace("/n", "")
        .replace(".", "")
        .replace(":", "")
        .replace(")", "")
        .replace("(", "")
        .replace(".", "")
        .replace("+", ""),
    flags=re.IGNORECASE)

def splitEmails(value):
    return re.split("[/,]+", str(value).replace(" ", ""), flags=re.IGNORECASE)
        
def generateNumbers(val):
    counter = 0
    number_list = list()
    stack = list()
    codes = list()
    numberstrings = list()
    data = splitNumbers(val)
    for s in range(0, 10):
        numberstrings.append(s)
    for code, area in LANDLINES.items():
        codes.append(str(code))
    for v in data:
        if(v != ''):
            if(len(v) >= 4):
                stack = v
                number_list.append(v)
            else:
                if(len(data) > 2):
                    stack = number_list[counter-1]
                rel = stack[0:len(stack)-len(v)]
                number_list.append(rel + v)
            counter += 1
            if(counter == len(data)):
                return number_list

def generateEmails(val):
    return splitEmails(val)

class DataSetup:
    def __init__():
        _brokers = dict()
        _warehouses = dict()
        _buyers = dict()
        _producers = dict()
        _producers_meta = dict()

    def init_brokers():
        bc = 0
        for row in data_brokers.values:
            inner_counter = 0
            inner = list()
            for value in row:
                if(inner_counter >= 3):
                    if(value != None):
                        if(value == "-" or value == "_" or value == "_'" or value == "'_"):
                            inner.append(None)
                        else:
                            inner.append(value)
                    else:
                        inner.append(None)
                inner_counter += 1
            if(bc >= 1):
                if(bc >= 2):
                    brokers[bc-2] = inner
            bc += 1
        for value in brokers:
            if(brokers[value][4] != None):
                brokers[value][4] = str(json.dumps(generateNumbers(brokers[value][4])))
            if(brokers[value][5] != None):
                brokers[value][5] = str(json.dumps(generateNumbers(brokers[value][5])))
            if(brokers[value][6] != None):
                brokers[value][6] = str(json.dumps(generateEmails(brokers[value][6])))

    def init_warehouses():
        wc = 0
        inner = list()
        for row in data_warehouses.values:
            inner_counter = 0
            inner = list()
            for value in row:
                if(inner_counter >= 3):
                    if(value != None):
                        if(value == "-" or value == "_" or value == "_'" or value == "'_"):
                            inner.append(None)
                        else:
                            inner.append(value)
                    else:
                        inner.append(None)
                inner_counter += 1
            if(wc >= 1):
                if(wc >= 2):
                    if(wc >= 83 and wc <= 85):
                        pass
                    else:
                        if(len(inner) >= 2 and inner != None):
                            inner.insert(1, inner[1])
                        warehouses[wc-2] = inner
            wc += 1
        for value in warehouses:
            if(warehouses[value][5] != None):
                warehouses[value][5] = str(json.dumps(generateNumbers(warehouses[value][5])))
            if(warehouses[value][6] != None):
                warehouses[value][6] = str(json.dumps(generateNumbers(warehouses[value][6])))
            if(warehouses[value][7] != None):
                warehouses[value][7] = str(json.dumps(generateEmails(warehouses[value][7])))

    def init_buyers():
        bc = 0
        inner = list()
        for row in data_buyers.values:
            inner_counter = 0
            inner = list()
            for value in row:
                if(inner_counter >= 3):
                    if(value != None):
                        if(value == "-" or value == "_" or value == "_'" or value == "'_"):
                            inner.append(None)
                        else:
                            inner.append(value)
                    else:
                        inner.append(None)
                inner_counter += 1
            if(bc >= 1):
                if(bc >= 2):
                    buyers[bc-2] = inner
            bc += 1
        for value in buyers:
            if(buyers[value][4] != None):
                buyers[value][4] = str(json.dumps(generateNumbers(buyers[value][4])))
            if(buyers[value][5] != None):
                buyers[value][5] = str(json.dumps(generateNumbers(buyers[value][5])))
            if(buyers[value][6] != None):
                buyers[value][6] = str(json.dumps(generateEmails(buyers[value][6])))
    
    def init_producers():
        pc = 0
        inner = list()
        for row in data_producers.values:
            inner_counter = 0
            inner = list()
            for value in row:
                if(inner_counter <= 5):
                    if(inner_counter == 1):
                        if(value != None and value != 'CODE'):
                            producers_relation.append(value)
                    if(value != None):
                        inner.append(value)
                    else:
                        inner.append(None)
                inner_counter += 1
            if(pc >= 1):
                if(pc >= 1):
                    producers[pc-1] = inner
            pc += 1

    def init_producers_meta():
        bc = 0
        inner = list()
        for row in data_producers_meta.values:
            inner_counter = 0
            inner = list()
            for value in row:
                if(inner_counter >= 4):
                    if(value != None):
                        if(value == "-" or value == "_" or value == "_'" or value == "'_"):
                            inner.append(None)
                        else:
                            inner.append(value)
                    else:
                        inner.append(None)
                inner_counter += 1
            if(bc >= 1):
                if(bc >= 3):
                    producers_meta[bc-3] = inner
            bc += 1
        for value in producers_meta:
            if(producers_meta[value][4] != None):
                producers_meta[value][4] = generateNumbers(producers_meta[value][4])
            if(producers_meta[value][5] != None):
                producers_meta[value][5] = generateNumbers(producers_meta[value][5])
            if(producers_meta[value][6] != None):
                producers_meta[value][6] = generateEmails(producers_meta[value][6])

    def migrate_brokers():
        multiinsert = list()
        for val in brokers:
            ins = tuple(brokers[val])
            multiinsert.append(ins)
        try:
            with connect(**CONNECTOR) as connection:
                init_values = '''INSERT INTO `brokers`
                (`company`, `code`, `postal_address`, `location`, `telephone`, `fax`, `email_address`)
                VALUES (%s, %s, %s, %s, %s, %s, %s)'''
                with connection.cursor() as cursor:
                    cursor.executemany(init_values, multiinsert)
                    connection.commit()
        except Error as e:
            print(e)
    
    def migrate_warehouses():
        multiinsert = list()
        for val in warehouses:
            ins = tuple(warehouses[val])
            multiinsert.append(ins)
        try:
            with connect(**CONNECTOR) as connection:
                init_values = '''INSERT INTO `warehouses`
                (`company`, `company_parent`, `code`, `postal_address`, `location`, `telephone`, `fax`, `email_address`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)'''
                with connection.cursor() as cursor:
                    cursor.executemany(init_values, multiinsert)
                    connection.commit()
        except Error as e:
            print(e)
    
    def migrate_buyers():
        multiinsert = list()
        for val in buyers:
            ins = tuple(buyers[val])
            multiinsert.append(ins)
        try:
            with connect(**CONNECTOR) as connection:
                init_values = '''INSERT INTO `buyers`
                (`company`, `code`, `postal_address`, `location`, `telephone`, `fax`, `email_address`)
                VALUES (%s, %s, %s, %s, %s, %s, %s)'''
                with connection.cursor() as cursor:
                    cursor.executemany(init_values, multiinsert)
                    connection.commit()
        except Error as e:
            print(e)
    
    def migrate_producers():
        multiinsert = list()
        for val in producers:
            ins = tuple(producers[val])
            multiinsert.append(ins)
        try:
            with connect(**CONNECTOR) as connection:
                init_values = '''INSERT INTO `producers`
                (`company`, `code`, `mark`, `country`, `warehouse`, `warehouse_code`)
                VALUES (%s, %s, %s, %s, %s, %s)'''
                with connection.cursor() as cursor:
                    cursor.executemany(init_values, multiinsert)
                    connection.commit()
        except Error as e:
            print(e)
            
    def migrate_producers_meta():
        for val in producers_meta:
            ins = tuple(producers_meta[val])
        try:
            with connect(**CONNECTOR) as connection:
                generator = list()
                counter = 0
                for row in producers_meta:
                    inner_counter = 0
                    for value in producers_meta[counter]:
                        clause = list()
                        if(inner_counter == 1):
                            code = value
                        if(inner_counter == 2):
                            postal_address = value
                        if(inner_counter == 3):
                            location = value
                        if(inner_counter == 4):
                            if(value != None):
                                telephone = str(json.dumps(value))
                            else:
                                telephone = None
                        if(inner_counter == 5):
                            if(value != None):
                                fax = str(json.dumps(value))
                            else:
                                fax = None
                        if(inner_counter == 6):
                            if(value != None):
                                email_address = str(json.dumps(value))
                            else:
                                email_address = None
                        inner_counter += 1
                    clause = [postal_address, location, telephone, fax, email_address, code]
                    generator.append(clause)
                    counter += 1
                generator = tuple(generator)
                init_values = 'UPDATE `producers` SET `postal_address` = %s, `location` = %s, `telephone` = %s, `fax` = %s, `email_address` = %s WHERE `code` = %s'
                with connection.cursor() as cursor:
                    cursor.executemany(init_values, generator)
                    connection.commit()
        except Error as e:
            print(e)

# -- INIT DATA --
DataSetup.init_warehouses()
DataSetup.init_brokers()
DataSetup.init_buyers()
DataSetup.init_producers()
DataSetup.init_producers_meta()
# -- INIT DATA --
# --
# -- MIGRATE DATA --
def DatabaseMigrate():
    DataSetup.migrate_warehouses()
    print('-- INITIALIZING WAREHOUSES --')
    DataSetup.migrate_brokers()
    print('-- INITIALIZING BROKERS --')
    DataSetup.migrate_buyers()
    print('-- INITIALIZING BUYERS --')
    DataSetup.migrate_producers()
    print('-- INITIALIZING PRODUCERS --')
    def updateMetaDB():
        DataSetup.migrate_producers_meta()
        print('-- INITIALIZING PRODUCERS META DATA --')
    t = Timer(4.0, updateMetaDB)
    t.start()
    def finalizeDB():
        print('-- FINALIZING DATABASE VALUES --')
    fd = Timer(6.0, finalizeDB)
    fd.start()
    def completeDB():
        print('-- DATABASE VALUES READY --')
    cd = Timer(10.0, completeDB)
    cd.start()
# -- MIGRATE DATA --
# --
# ------------------------------------------ CRITICAL DATA : EXECUTE ONLY ONCE
# DatabaseMigrate()
# ------------------------------------------ CRITICAL DATA : EXECUTE ONLY ONCE