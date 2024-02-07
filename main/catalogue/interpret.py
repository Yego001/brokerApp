from openpyxl import load_workbook, Workbook
import pandas as pd
from mysql.connector import connect, Error
import json
from django.core.files.storage import FileSystemStorage
from main.catalogue.connector import CONNECTOR
from main.catalogue.format import *
from main.catalogue.helper import Helper
from main.catalogue.query import *
import re
from threading import Timer, Thread
from queue import Queue
from datetime import datetime
from django.utils.dateformat import DateFormat
from django.core.cache import cache
import math

cache.clear()

def LotQuery():
    try:
        with connect(**CONNECTOR) as connection:
            current_lot = '''SELECT `number` FROM `lot_number`'''
            with connection.cursor() as cursor:
                cursor.execute(current_lot)
                row = cursor.fetchone()
                return row[0]
                        
    except Error as e:
            print(e)

LOT_COUNTER = LotQuery()
def LotCounter():
    global LOT_COUNTER
    if(LOT_COUNTER < 25999):
        LOT_COUNTER += 1
    else:
        LOT_COUNTER = 23000
    return LOT_COUNTER

def CloseLot(lot_number):
    try:
        with connect(**CONNECTOR) as connection:
            update_lot = '''UPDATE `lot_number` SET `number` = %s'''
            with connection.cursor() as cursor:
                cursor.execute(update_lot, (lot_number,))
                connection.commit()
                        
    except Error as e:
            print(e)
            
TEAGRADES_RELATION = [
    'type',
    'class',
    'grade',
    'packages',
    'package_type'
]
    
def TeaGradesQuery():
    try:
        with connect(**CONNECTOR) as connection:
            current_lot = '''SELECT `type`, `class`, `grade`, `packages`, `package_type` FROM `teagrades`'''
            with connection.cursor() as cursor:
                cursor.execute(current_lot)
                data = cursor.fetchall()
                parsed_data = list()
                for value in data:
                    inner_counter = 0
                    field_values = dict()
                    for field in value:
                        if(TEAGRADES_RELATION[inner_counter] == 'packages' or TEAGRADES_RELATION[inner_counter] == 'package_type'):
                            if(field != None):
                                field_values[TEAGRADES_RELATION[inner_counter]] = json.loads(field)
                            else:
                                field_values[TEAGRADES_RELATION[inner_counter]] = None
                        else:
                            field_values[TEAGRADES_RELATION[inner_counter]] = field
                        inner_counter += 1
                    parsed_data.append(field_values)
                return parsed_data
                        
    except Error as e:
            print(e)
            
DATA_COLUMNS = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','R','S','T','U','V','W','X','Y','Z']

DATA_ALIAS = {
    "company": ["company", "company code"],
    "invoice_number": ["invoice", "invoice number", "invoice no", "invoice no."],
    "warehouse": ["warehouse", "warehouse location"],
    "grade": ["grade"],
    "mark": ["mark", "garden"],
    "gross": ["gross weight", "gross wt", "gross", "grossweight", "total gross wt(kgs)"],
    "net": ["net", "net weight", "total net weight", "net wt", "weight", "netweight", "netweight(in kg)", "total net wt (kgs)"],
    "type": ["type", "packing type", "qty", "packaging", "packaging type"],
    # "bags": ["no of bags", "quantity", "qty", "pkgs", "packages"],
    "packages": ["no of bags", "quantity", "qty", "pkgs", "packages"],
    "manufacture_date": ["date manufacture", "mfg. date", "manufacture date"],
    "packed_date": ["date packed", "date"],
    "sample_weight": ["sample weight", "sample wt", "sample"],
    "pallet_weight": ["pallet weight", "pallet wt", "pallet"],
    "RA": ["ra"],
    "tare": ["tare", "tare weight", "tareweight(in kg)", "tare weight(in kg)"],
    "kgs": ["kgs", "weight per bag"],
    "reprint": ["reprint", "reprints"],
    "ignore": ["", None],
}

DATA_CATALOGUE = {
    "comments": "Comments",
    "empty1": "",
    "warehouse": "Ware Hse",
    "entry_number": "Entry No.",
    "value": "Value",
    "empty2": "",
    "lot": "Lot No.",
    "company": "Company",
    "mark": "Mark",
    "grade": "Grade",
    "manufacture_date": "Manf Date",
    "RA": "RA",
    "empty3": "",
    "invoice_number": "Invoice",
    "packages": "Pkgs",
    "type": "Type",
    "net": "Net",
    "gross": "Gross",
    "kgs": "Kgs",
    "tare": "Tare",
    "sale_price": "Sale Price",
    "buyers_and_packages": "Buyer & Packages",
    "bags": "No. of Bags",
    "reprint": "REPRINT",
}

TEAGRADES_ORDER = [
    'BP1',
    'PF1',
    'PD',
    'D1',
    'DUST1',
    'BP',
    'PF',
    'PF2',
    'FNGS1',
    'FNGS',
    'DUST',
    'DUST2',
    'BMF',
    'BMF1',
]

TEAGRADES_DATA = dict()
TEAGRADES_LIST = list()
for val in TeaGradesQuery():
    TEAGRADES_DATA[val['grade']] = val['type']
    TEAGRADES_LIST.append(val['grade'])
TEAGRADES_DUST = ['PD', 'D1', 'D1 ', 'DUST1']

def findActualGrade(grade):
    if(grade in TEAGRADES_DUST):
        return 'dust'
    else:
        return TEAGRADES_DATA[grade]

def convertFloat(val):
    if(val != None):
        return float(val)
    else:
        return None

class MetaLogic():
    # tare = gross - net
    # gross = net + tare
    # net = gross + tare
    def _tare(gross, net):
        return int(gross) - int(net)
    def _gross(net, tare):
        return int(net) + int(tare)
    def _net(gross, tare):
        return int(gross) + int(tare)
    # kgs = net / bags
    # net = kgs * bags
    # bags = kgs * net
    def _kgs(net, bags):
        return convertFloat(net) / convertFloat(bags)
    def _kgs_alt(pkgs, net):
        return convertFloat(net) / convertFloat(pkgs)
    def _net_alt(kgs, bags):
        return int(kgs) * int(bags)
    def _bags(kgs, net):
        return int(kgs) * int(net)

def findAlias(list, value):
    perfect = [
        'kgs',
        'pkgs',
        'ra'
    ]
    end = len(list)
    counter = 1
    for alias in list:
        if(value in perfect):
            if(re.search(alias, value) != None):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        else:
            if(value == alias):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        counter += 1
    

def getAliasRelation(value):
    endif = 0
    for alias in DATA_ALIAS:
        endif += 1
        if(findAlias(DATA_ALIAS[alias], str(value).lower())):
            return alias
        else:
            if endif == len(DATA_ALIAS):
                return alias
            else:
                continue

COLUMN_ASSOC = dict()
class DataInterpretor:
    def init_data(left_bound, data_layer, file):
        left_bound = int(left_bound)
        data_layer = int(data_layer)
        
        folder='media/documents/allocations/'
        fs = FileSystemStorage(location=folder)
        file_data = fs.open(file, 'rb')
        
        WORKBOOK = load_workbook(filename = file_data, data_only=True)
        file_data.close()
        DATA = {}
        RELATION = []
        sheet = WORKBOOK.worksheets[0]
        bc = 0
        inner = list()
        endif_check = list()
        for row in sheet.values:
            allow = True
            inner_counter = 0
            for value in row:
                if(inner_counter >= left_bound-1):
                    if(value != None):
                        if(bc == data_layer-1):
                            # value = re.sub(r'\t', '', value)
                            RELATION.append(value)
                        # COLUMN_ASSOC[bc] = DATA_COLUMNS[bc]
                        inner.append(value)
                    else:
                        if(bc == data_layer-1):
                            RELATION.append("ignore")
                        inner.append(None)
                        endif_check.append(value)
                inner_counter += 1
            if(len(endif_check) >= 5 and bc >= 5):
                allow = False
                break
            elif(bc >= data_layer and len(endif_check) < 5):
                if allow: DATA[bc-(data_layer)] = row[slice(left_bound-1, len(row))]
            inner = list()
            endif_check = list()
            bc += 1
        return {
            'relation': RELATION,
            'data': DATA
        }

    def generate_data(data):
        set_data = list()
        counter = 0
        for vals in data['data'].values():
            inner_counter = 0
            inner_dict = dict()
            for inner_data in vals:
                relation = getAliasRelation(data['relation'][inner_counter])
                if(relation != False):
                    inner_dict[relation] = str(inner_data)
                else:
                    inner_dict[relation] = None
                inner_counter += 1
            set_data.append(inner_dict)
            counter += 1
        return set_data

STACK_DATA = {}
combined = list()
populated = list()
POPULATED_MAIN = list()
POPULATED_DUST = list()
POPULATED_SECONDARY = list()
CATALOGUE_KEYS = list()
UNIQUE_MARKS = list()
catalogue_main_leafy = None
catalogue_main_dust = None
catalogue_secondary = None
catalogue_valuations = None
# ---- PREDEFINED
SHEETS_ORDER = ['primary', 'secondary']
RELATION_ORDER = ['primary', 'dust', 'secondary']

def RelationshipSort(obj, rel, key):
    sorted_object = list()
    for item in rel:
        for val in obj:
            for _key in val:
                if _key == key:
                    value = val[_key]
                    if value == item:
                        sorted_object.append(val)
    return sorted_object

def generate_sorter(data_val):
    rel = list()
    sorted_rel = list()
    if len(data_val) > 0:
        for item in data_val:
            _val = str(item["invoice_number"]).split('||')[0]
            _val3 = None
            _type = "str"
            _type3 = "str"
            if _val.isdigit():
                rel.append(int(_val))
                partition = 1
                _type = "int"
            else:
                rel.append(_val)
                partition = len(rel[0].split('/'))
                if partition == 3:
                    _val3 = item["invoice_number"].split('||')[0].split('/')[2]
                    if _val3.isdigit():
                        _type3 = "int"
        if partition == 1:
            if _type == "int":
                data_val.sort(key = lambda i: int(str(i.get("invoice_number")).split('||')[0]))
            else:
                data_val.sort(key = lambda i: i.get("invoice_number").split('||')[0])
            sorted_rel = data_val
        elif partition == 3:
            if _type3 == "int":
                data_val.sort(key = lambda i: int(i.get("invoice_number").split('||')[0].split('/')[2]))
            else:
                data_val.sort(key = lambda i: i.get("invoice_number").split('||')[0])
            sorted_rel = data_val
        elif partition == 2:
            partition_data = list()
            ys = list()
            ys_data = dict()
            ys_data_retain = dict()
            sorted_unpartition = list()
            for item in rel:
                partition_data.append(
                    {
                        "n":item.split('/')[0],
                        "y": item.split('/')[1],
                    }
                )
                ys.append(item.split('/')[1])
            l = set()
            ys =  [x for x in ys if x not in l and l.add(x) is None]
            ys.sort()
            for y in ys:
                ys_data[y] = []
            sample_val = "str"
            for yr in ys_data:
                for data in partition_data:
                    if data['y'] == yr:
                        if data['n'].isdigit():
                            sample_val = "int"
                        if sample_val == "int":
                            ys_data[yr].append(int(data['n']))
                            ys_data_retain[int(data['n'])] = data['n']
                        else:
                            ys_data[yr].append(data['n'])
                ys_data[yr].sort()
            for yr in ys_data:
                for data in ys_data[yr]:
                    if sample_val == "int":
                        sorted_unpartition.append(f'{ys_data_retain[data]}/{yr}')
                    else:
                        sorted_unpartition.append(f'{data}/{yr}')
            sorted_rel = sorted(data_val, key=lambda x:sorted_unpartition.index(x['invoice_number'].split('||')[0]))
        else:
            data_val.sort(key = lambda i: i.get("invoice_number").split('||')[0])
            sorted_rel = data_val
    else:
        sorted_rel = data_val
    return sorted_rel

def StackGenerator(input_data):
    global STACK_DATA
    global combined
    counter = 0
    for file in input_data:
        STACK_DATA[counter] = DataInterpretor.generate_data(
            DataInterpretor.init_data(
                file['left_bound'],
                file['data_layer'],
                file['file']
            )
        )
        counter += 1
    for data in STACK_DATA.values():
        combined += data
    for data in combined:
        exist = list()
        inner_val = dict()
        for single in data:
            if(single in DATA_CATALOGUE.keys()):
                exist.append(single)
        for value in DATA_CATALOGUE.keys():
            if(value in exist):
                if value == "invoice_number":
                    inner_val[value] = f'{str(data[value])}||{str(data["mark"])}'
                else:
                    inner_val[value] = data[value]
            else:
                inner_val[value] = None
        populated.append(inner_val)
        
INVOICE_GRADE_RELATIONS = dict()
INVOICE_MARK_RELATIONS = dict()
INVOICE_SORTED_GRADE = {}
INVOICE_SORTED_MARK = {}
GRADES_PRIMARY = dict()
GRADES_DUST = dict()
GRADES_SECONDARY = dict()
INVOICE_ORDERS = dict()
initial_sorter = {
    'primary': {},
    'secondary': {}
}
initial_sorter_mark = {
    'primary': {},
    'secondary': {}
}
# ----- POPULATED DATA
POPULATED_MAIN_SORTED = list()
POPULATED_DUST_SORTED = list()
POPULATED_SECONDARY_SORTED = list()
DATA_POPULATED = dict()
DATA_POPULATED_GROUPED = dict()
DATA_POPULATED_GRADED = dict()
DATA_POPULATED_UNIQUE_GRADES = dict()
# GLOBAL DATA
GLOBAL_DATA_POPULATED = dict()
GLOBAL_DATA_SECONDARY = list()
GLOBAL_DATA_PRIMARY = list()
GLOBAL_DATA_POPULATED_MAIN = list()
GLOBAL_DATA_POPULATED_DUST = list()
GLOBAL_DATA_POPULATED_SECONDARY = list()

catalogue = Workbook()
def PopulateInitialData(marks_order, reprint_placement):
    global POPULATED_MAIN
    global POPULATED_SECONDARY
    global POPULATED_DUST
    global CATALOGUE_KEYS
    global UNIQUE_MARKS
    global INVOICE_GRADE_RELATIONS
    global INVOICE_MARK_RELATIONS
    global INVOICE_SORTED_GRADE
    global INVOICE_SORTED_MARK
    global GRADES_PRIMARY
    global GRADES_DUST
    global GRADES_SECONDARY
    global INVOICE_ORDERS
    global catalogue_main_leafy
    global catalogue_main_dust
    global catalogue_secondary
    global catalogue_valuations
    global populated
    global DATA_POPULATED
    global DATA_POPULATED_GROUPED
    global DATA_POPULATED_GRADED
    global DATA_POPULATED_UNIQUE_GRADES
    
    global GLOBAL_DATA_POPULATED
    global GLOBAL_DATA_SECONDARY
    global GLOBAL_DATA_PRIMARY

    catalogue_main_leafy = catalogue.active
    catalogue_main_leafy.title = "Main_Leafy Catalogue"
    catalogue_main_dust = catalogue.create_sheet(title="Main_Dust")
    catalogue_secondary = catalogue.create_sheet(title="Secondary_Catalogue")
    catalogue_valuations = catalogue.create_sheet(title="Valuations")

    CATALOGUE_KEYS = list(DATA_CATALOGUE.keys())
    for row in range(1, 2):
        for col in range(1, 2):
            catalogue_main_leafy.cell(column=col, row=row, value="Prime Tea Brokers Limited")
            catalogue_main_dust.cell(column=col, row=row, value="Prime Tea Brokers Limited")
            catalogue_secondary.cell(column=col, row=row, value="Prime Tea Brokers Limited")
            catalogue_valuations.cell(column=col, row=row, value="PRIME TEA BROKERS LIMITED")

    for row in range(2, 3):
        for col in range(1, 2):
            catalogue_main_leafy.cell(column=col, row=row, value="Main Sale on this date")
            catalogue_main_dust.cell(column=col, row=row, value="Main Sale on this date")
            catalogue_secondary.cell(column=col, row=row, value="Secondary Sale on this date")
            catalogue_valuations.cell(column=col, row=row, value="Valuations for the Sale on this date")

    for row in range(3, 4):
        # bags include fallback: +1
        # reprint include fallback: +1
        # Total include fallback: +2
        for col in range(1, len(CATALOGUE_KEYS)-1):
            catalogue_main_leafy.cell(column=col, row=row, value=DATA_CATALOGUE[CATALOGUE_KEYS[col-1]])
            catalogue_main_dust.cell(column=col, row=row, value=DATA_CATALOGUE[CATALOGUE_KEYS[col-1]])
            catalogue_secondary.cell(column=col, row=row, value=DATA_CATALOGUE[CATALOGUE_KEYS[col-1]])

    Format.formatTitle(catalogue_main_leafy['A1'])
    Format.formatTitle(catalogue_main_dust['A1'])
    Format.formatTitle(catalogue_secondary['A1'])
    Format.formatTitle(catalogue_valuations['A1'])
    Format.formatSubTitle(catalogue_main_leafy['A2'])
    Format.formatSubTitle(catalogue_main_dust['A2'])
    Format.formatSubTitle(catalogue_secondary['A2'])
    Format.formatSubTitle(catalogue_valuations['A2'])
    
    catalogue_main_leafy.row_dimensions[1].height = 37
    catalogue_main_leafy.row_dimensions[2].height = 37
    catalogue_main_dust.row_dimensions[1].height = 37
    catalogue_main_dust.row_dimensions[2].height = 37
    catalogue_secondary.row_dimensions[1].height = 37
    catalogue_secondary.row_dimensions[2].height = 37
    catalogue_valuations.row_dimensions[1].height = 37
    catalogue_valuations.row_dimensions[2].height = 37
            
    catalogue_main_leafy.merge_cells('A1:W1')
    catalogue_main_leafy.merge_cells('A2:W2')
    catalogue_main_dust.merge_cells('A1:W1')
    catalogue_main_dust.merge_cells('A2:W2')
    catalogue_secondary.merge_cells('A1:W1')
    catalogue_secondary.merge_cells('A2:W2')
    catalogue_valuations.merge_cells('A1:W1')
    catalogue_valuations.merge_cells('A2:W2')
    
    populated = RelationshipSort(populated, TEAGRADES_ORDER, 'grade')
    
    for value in populated:
        for data in value:
            val = value[data]
            if(data == 'mark'):
                UNIQUE_MARKS.append(val)

    lookup = set()
    UNIQUE_MARKS = [x for x in UNIQUE_MARKS if x not in lookup and lookup.add(x) is None]
    UNIQUE_MARKS = marks_order
    # UNIQUE_MARKS = ['SIOMO', 'EMROK', 'TULON', 'SARMA', 'TEGAT', 'CUPATEA', 'KABIANGA', 'KIPSINENDE', 'SIAN']
        
    sales = 0
    for sale in populated:
        for value in sale:
            if(value == 'grade'):
                sale_value = Helper.remove_space(populated[sales][value])
                if(sale_value):
                    type = TEAGRADES_DATA[sale_value]
                else: type = None
                if(populated[sales]['grade']):
                    INVOICE_GRADE_RELATIONS[populated[sales]['invoice_number']] = populated[sales]['grade'].replace(' ', '')
                else:
                    INVOICE_GRADE_RELATIONS[populated[sales]['invoice_number']] = populated[sales]['grade']
                if(populated[sales]['mark']):
                    INVOICE_MARK_RELATIONS[populated[sales]['invoice_number']] = populated[sales]['mark'].replace(' ', '')
                else:
                    INVOICE_MARK_RELATIONS[populated[sales]['invoice_number']] = populated[sales]['mark']
                if(sale_value in TEAGRADES_DUST):
                    GRADES_DUST[populated[sales]['invoice_number']] = populated[sales]['mark']
                    POPULATED_DUST.append(sale)
                else:
                    if(type == 'primary'):
                        GRADES_PRIMARY[populated[sales]['invoice_number']] = populated[sales]['mark']
                        POPULATED_MAIN.append(sale)
                    elif(type == 'secondary'):
                        GRADES_SECONDARY[populated[sales]['invoice_number']] = populated[sales]['mark']
                        POPULATED_SECONDARY.append(sale)
                    else:
                        POPULATED_MAIN.append(sale)
                # Replace Bounds [ MARK-ed invoice_number ]
        sales += 1
        
    DATA_POPULATED = {
        'primary': POPULATED_MAIN,
        'secondary': POPULATED_SECONDARY,
    }
    
    DATA_POPULATED_GROUPED = {
        'primary': {},
        'secondary': {},
    }
    
    DATA_POPULATED_GRADED = {
        'primary': {},
        'secondary': {},
    }
    
    GLOBAL_DATA_POPULATED = {
        'primary': [],
        'secondary': [],
    }
    
    for group in DATA_POPULATED_GROUPED:
        for mark in UNIQUE_MARKS:
            DATA_POPULATED_GROUPED[group][mark] = list()
            DATA_POPULATED_GRADED[group][mark] = list()
    
    for type in DATA_POPULATED_GROUPED:
        for mark in DATA_POPULATED_GROUPED[type]:
            unique_getter = list()
            marked_initialize = dict()
            for sale in populated:
                for value in sale:
                    if(value == 'mark'):
                        if sale[value] == mark:
                            grade_type = TEAGRADES_DATA[sale['grade']]
                            if grade_type == type:
                                DATA_POPULATED_GROUPED[type][mark].append(sale)
                                unique_getter.append(sale['grade'])
            lookup = set()
            sorted_unique_getter = [x for x in unique_getter if x not in lookup and lookup.add(x) is None]
            for grade in sorted_unique_getter:
                marked_initialize[grade] = []
            DATA_POPULATED_GRADED[type][mark] = marked_initialize

    for type in DATA_POPULATED_GROUPED:
        for mark in DATA_POPULATED_GROUPED[type]:
            DATA_POPULATED_GROUPED[type][mark] = RelationshipSort(DATA_POPULATED_GROUPED[type][mark], TEAGRADES_ORDER, 'grade')

    # Global Grade Order - ignore reprints
    for type in DATA_POPULATED_GROUPED:
        counter = 0
        for mark in DATA_POPULATED_GROUPED[type]:
            for grade in DATA_POPULATED_GRADED[type][mark]:
                for sale in DATA_POPULATED_GROUPED[type][mark]:
                    for value in sale:
                        if(value == 'grade'):
                            if sale[value] == grade:
                                if sale['reprint'] == None or sale['reprint'] == 'None':
                                    DATA_POPULATED_GRADED[type][mark][grade].append(sale)
                    counter += 1
                    
    _GLOBAL_DATA = [*POPULATED_MAIN, *POPULATED_SECONDARY]
    _REPRINTS = list()
    for lot in _GLOBAL_DATA:
        for value in lot:
            if value == 'reprint':
                if lot[value] != None:
                    _REPRINTS.append(lot)
    
    for type in DATA_POPULATED_GRADED:
        for mark in DATA_POPULATED_GRADED[type]:
            for grade in DATA_POPULATED_GRADED[type][mark]:
                DATA_POPULATED_GRADED[type][mark][grade] = generate_sorter(DATA_POPULATED_GRADED[type][mark][grade])
                # Delegate sorting to function
                # DATA_POPULATED_GRADED[type][mark][grade].sort(key=lambda item: item.get("invoice_number"))
                
    # Secondary Grade Order - reprints placement
    for type in DATA_POPULATED_GROUPED:
        counter = 0
        for mark in DATA_POPULATED_GROUPED[type]:
            for grade in DATA_POPULATED_GRADED[type][mark]:
                reprint_list = list()
                for sale in DATA_POPULATED_GROUPED[type][mark]:
                    for value in sale:
                        if(value == 'grade'):
                            if sale[value] == grade:
                                if sale['reprint'] != None and sale['reprint'] != 'None':
                                    ##### - Resolve unsorted/reverse-sorted invoice numbers
                                    reprint_list.append(sale)
                                    # Start Resolution
                                    # if reprint_placement == 'first':
                                    #     DATA_POPULATED_GRADED[type][mark][grade].insert(0, sale)
                                    # elif reprint_placement == 'last':
                                    #     DATA_POPULATED_GRADED[type][mark][grade].append(sale)
                                    # End Resolution
                    counter += 1
                # Resolve unsorted/reverse-sorted invoice numbers ###
                # Add as single list rather than append item by item ###
                if len(reprint_list) > 0:
                    reprint_list = generate_sorter(reprint_list)
                    # Delegate sorting to function
                    # reprint_list.sort(key=lambda item: item.get("invoice_number"))
                    # SORT reprints by invoice number
                    if reprint_placement == 'first':
                        DATA_POPULATED_GRADED[type][mark][grade] = [*reprint_list, *DATA_POPULATED_GRADED[type][mark][grade]]
                    elif reprint_placement == 'last':
                        DATA_POPULATED_GRADED[type][mark][grade] = [*DATA_POPULATED_GRADED[type][mark][grade], *reprint_list]

    for type in DATA_POPULATED_GRADED:
        for mark in DATA_POPULATED_GRADED[type]:
            for grade in DATA_POPULATED_GRADED[type][mark]:
                for sale in DATA_POPULATED_GRADED[type][mark][grade]:
                    GLOBAL_DATA_POPULATED[type].append(sale)

    GLOBAL_DATA_PRIMARY = GLOBAL_DATA_POPULATED['primary']
    GLOBAL_DATA_SECONDARY = GLOBAL_DATA_POPULATED['secondary']

    for mark in UNIQUE_MARKS:
        initial_sorter['primary'][mark] = {}
        initial_sorter_mark['primary'][mark] = {}

def AssignLotNumbers(lotnumber):
    lotvalue = int(lotnumber)
    counter = 0
    for sale in GLOBAL_DATA_PRIMARY:
        for value in sale:
            if value == 'lot':
                GLOBAL_DATA_PRIMARY[counter][value] = lotvalue
        counter += 1
        if lotvalue < 25999:
            lotvalue += 1
        else: lotvalue = 23000
    counter = 0
    for sale in GLOBAL_DATA_SECONDARY:
        for value in sale:
            if value == 'lot':
                GLOBAL_DATA_SECONDARY[counter][value] = lotvalue
        counter += 1
        if lotvalue < 25999:
            lotvalue += 1
        else: lotvalue = 23000
    return lotvalue

def RowSeparator(company, mark, ra):
    if ra and ra != None and ra != "None":
        ra = '(%s Certified)'%ra
    return {
        'comments': None,
        'empty1': None,
        'warehouse': company,
        'entry_number': None,
        'value': None,
        'empty2': None,
        'lot': mark,
        'company': None,
        'mark': None,
        'grade': None,
        'manufacture_date': None,
        'RA': None,
        'empty3': None,
        'invoice_number': None,
        'packages': None,
        'type': ra,
        'net': None,
        'gross': None,
        'kgs': None,
        'tare': None,
        'sale_price': None,
        'buyers_and_packages': None,
        'bags': None,
        'reprint': None,
    }
    
def SeparatorPopulator(data):
    ACCESS_LIST = list()
    POP_DATA = list()
    ACCESS_LIST.append(4)
    listen = data[0]['company']
    POP_DATA.append(
        RowSeparator(
            data[0]['company'],
            data[0]['mark'],
            data[0]['RA']
        )
    )
    counter = 0
    access_counter = 4
    for sale in data:
        if sale['mark'] != listen and counter != 0 and access_counter > 4:
            POP_DATA.append(
                RowSeparator(
                    data[counter]['company'],
                    data[counter]['mark'],
                    data[counter]['RA']
                )
            )
            access_counter += 1
            ACCESS_LIST.append(access_counter)
        POP_DATA.append(sale)
        listen = sale['mark']
        counter += 1
        access_counter += 1
    return {
        'DATA': POP_DATA,
        'ACCESS': ACCESS_LIST
    }

def InitGenerator():
    global GLOBAL_DATA_POPULATED_MAIN
    global GLOBAL_DATA_POPULATED_DUST
    global GLOBAL_DATA_POPULATED_SECONDARY
    
    for sale in GLOBAL_DATA_PRIMARY:
        for value in sale:
            if value == 'grade':
                if sale[value] in TEAGRADES_DUST:
                    GLOBAL_DATA_POPULATED_DUST.append(sale)
                else:
                    GLOBAL_DATA_POPULATED_MAIN.append(sale)
    
    GLOBAL_DATA_POPULATED_SECONDARY = GLOBAL_DATA_SECONDARY
    
    GLOBAL_DATA_MAIN_ACCESS = SeparatorPopulator(GLOBAL_DATA_POPULATED_MAIN)
    GLOBAL_DATA_DUST_ACCESS = SeparatorPopulator(GLOBAL_DATA_POPULATED_DUST)
    GLOBAL_DATA_SECONDARY_ACCESS = SeparatorPopulator(GLOBAL_DATA_POPULATED_SECONDARY)
        
    GLOBAL_DATA_POPULATED_MAIN = GLOBAL_DATA_MAIN_ACCESS['DATA']
    GLOBAL_DATA_POPULATED_DUST = GLOBAL_DATA_DUST_ACCESS['DATA']
    GLOBAL_DATA_POPULATED_SECONDARY = GLOBAL_DATA_SECONDARY_ACCESS['DATA']
        
    return {
        'MAIN_ACCESS': GLOBAL_DATA_MAIN_ACCESS['ACCESS'],
        'DUST_ACCESS': GLOBAL_DATA_DUST_ACCESS['ACCESS'],
        'SECONDARY_ACCESS': GLOBAL_DATA_SECONDARY_ACCESS['ACCESS'],
    }

def GenerateLot():
    pc = [*GLOBAL_DATA_PRIMARY, *GLOBAL_DATA_SECONDARY]
    for sale in pc:
        INVOICE_ORDERS[str(sale['invoice_number'])] = sale['lot']

def CumulativePopulate(lotnum, marks_order, reprint_placement):
    PopulateInitialData(marks_order, reprint_placement)
    lot = AssignLotNumbers(lotnum)
    CloseLot(LotCounter())
    separator_access = InitGenerator()
    GenerateLot()
    return {
        'lotnumber': lot,
        'catalogue_data': [*GLOBAL_DATA_PRIMARY, *GLOBAL_DATA_SECONDARY],
        'invoice_data': INVOICE_ORDERS,
        'account_sale_data': INVOICE_ORDERS,
        'separator_access': separator_access,
    }

def get_rem(n):
    low = math.floor(n/8)
    abs = math.ceil(n/8)
    rem = n%8
    return {
        "rem": rem,
        "rows": abs,
        "low": low,
    }

def row_arrange(l):
    col1 = list()
    col2 = list()
    col3 = list()
    col4 = list()
    col5 = list()
    col6 = list()
    col7 = list()
    col8 = list()
    cols = [col1, col2, col3, col4, col5, col6, col7, col8]
    data = get_rem(len(l))
    counter = 0
    splice_point = 0
    for col in cols:
        if(counter < data["rem"]):
            cols[counter] = l[splice_point:data["rows"]+splice_point]
            splice_point += data["rows"]
        elif(counter >= data["rem"]):
            cols[counter] = l[splice_point:data["low"]+splice_point]
            splice_point += data["low"]
        counter += 1
    datam = list(zip(cols[0],cols[1],cols[2],cols[3],cols[4],cols[5],cols[6],cols[7]))
    _data = list()
    c = 0
    for col in datam:
        _data = [*_data, *list(col)]
        c += 1
    for col in cols:
        if len(col) == data["rows"] and data["rows"] != data["low"]:
            _data.append(col[data["rows"]-1])
    return _data

def ArrangeLots(catalogue_data):
    VALUATIONS_LIST_MAIN = list()
    VALUATIONS_LIST_SECONDARY = list()
    VALUATIONS_ROW_LIST_MAIN = list()
    VALUATIONS_ROW_LIST_SECONDARY = list()
    all_lots = list()
    lots_list_main = list()
    lots_list_secondary = list()
    for sale in catalogue_data:
        for value in sale:
            if value == 'lot':
                all_lots.append(sale['lot'])
                
    lots_list = sorted(all_lots)
    def access_lot(lot):
        for sale in catalogue_data:
            for value in sale:
                if value == 'lot':
                    if sale[value] == lot:
                        type = TEAGRADES_DATA[sale['grade']]
                        sale_price = sale['value']
                        return {
                            "type": type,
                            "sale_price": sale_price
                        }
                        
    for lot in lots_list:
        type = access_lot(lot)["type"]
        if type == 'primary':
            lots_list_main.append(lot)
        else:
            lots_list_secondary.append(lot)
            
    lots_list_main = row_arrange(sorted(lots_list_main))
    lots_list_secondary = row_arrange(sorted(lots_list_secondary))
    for lot in lots_list_main:
        sale_price = access_lot(lot)["sale_price"]
        VALUATIONS_LIST_MAIN.append({
            'lot': lot,
            'sale_price': sale_price,
            'empty': None,
        })
    for lot in lots_list_secondary:
        sale_price = access_lot(lot)["sale_price"]
        VALUATIONS_LIST_SECONDARY.append({
            'lot': lot,
            'sale_price': sale_price,
            'empty': None,
        })
    counter = 0
    gcounter = 0
    inner = list()
    for sale in VALUATIONS_LIST_MAIN:
        inner.append(sale['lot'])
        inner.append(sale['sale_price'])
        inner.append(sale['empty'])
        counter += 1
        if counter == 8 or gcounter == len(VALUATIONS_LIST_MAIN)-1:
            counter = 0
            VALUATIONS_ROW_LIST_MAIN.append(inner)
            inner = list()
        gcounter += 1
    counter = 0
    gcounter = 0
    inner = list()
    for sale in VALUATIONS_LIST_SECONDARY:
        inner.append(sale['lot'])
        inner.append(sale['sale_price'])
        inner.append(sale['empty'])
        counter += 1
        if counter == 8 or gcounter == len(VALUATIONS_LIST_SECONDARY)-1:
            counter = 0
            VALUATIONS_ROW_LIST_SECONDARY.append(inner)
            inner = list()
        gcounter += 1
    return {
        'main': VALUATIONS_ROW_LIST_MAIN,
        'secondary': VALUATIONS_ROW_LIST_SECONDARY,
    }

class PopulateValuations:
    def __init__():
        return
    def fill(valuations):
        valuations_title = ['Lot No', 'USC', None]
        valuations_titlerow = list()
        for i in range(1, 9):
            valuations_titlerow = [*valuations_titlerow, *valuations_title]
        main_valuations = valuations['main']
        secondary_valuations = valuations['secondary']
        center_align = Alignment(
            horizontal='center',
            vertical='center'
        )
        
        # row 4 -- title
        for row in range(4, 5):
            for col in range(1, len(valuations_titlerow)):
                catalogue_valuations.cell(column=col, row=row, value=valuations_titlerow[col-1])
        # row 7 -- main
        for row in range(7, 8):
            for col in range(1, len(valuations_titlerow)):
                catalogue_valuations.cell(column=col, row=row, value='MAIN')
        catalogue_valuations.merge_cells('A7:W7')
        
        access_start = 8
        access_end = len(main_valuations)+access_start
        # row 8 -- main valuations
        for row in range(access_start, access_end):
            for col in range(1, len(main_valuations[row-access_start])):
                col_data = main_valuations[row-access_start][col-1]
                catalogue_valuations.cell(column=col, row=row, value=col_data)
                
        # dynamic row -- secondary
        for row in range(access_end, access_end+1):
            for col in range(1, len(valuations_titlerow)):
                catalogue_valuations.cell(column=col, row=row, value='SECONDARY')
        catalogue_valuations.merge_cells('A' + str(access_end) + ':W' + str(access_end))

        # FIXME: Add merge for primary and secondary titles
                
        secondary_row = access_end+1
        secondary_end = len(secondary_valuations)+secondary_row
        
        # dynamic row -- secondary valuations
        for row in range(secondary_row, secondary_end):
            for col in range(1, len(secondary_valuations[row-secondary_row])):
                col_data = secondary_valuations[row-secondary_row][col-1]
                catalogue_valuations.cell(column=col, row=row, value=col_data)
                
        zero_width = ['C','F','I','L','O','R','U']
                
        for column in zero_width:
            catalogue_valuations.column_dimensions[column].width = 1.8
                
    # column_widths = []
    # for row in data:
    #     for i, cell in enumerate(row):
    #         if len(column_widths) > i:
    #             if len(cell) > column_widths[i]:
    #                 column_widths[i] = len(cell)
    #         else:
    #             column_widths += [len(cell)]
        
    # for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
    #     worksheet.column_dimensions[get_column_letter(i)].width = column_width
    
    # for i in range(1, ws.max_column+1):
    #     ws.column_dimensions[get_column_letter(i)].bestFit = True
    #     ws.column_dimensions[get_column_letter(i)].auto_size = True
    
    # sheet.column_dimensions["B"].width = 50
    
class PopulateCatalogue:
    def fill(target, separator_access, marks_order):
        if(target == 'primary'):
            focus = GLOBAL_DATA_POPULATED_MAIN
            target_sheet = catalogue_main_leafy
        elif(target == 'dust'):
            focus = GLOBAL_DATA_POPULATED_DUST
            target_sheet = catalogue_main_dust
        elif(target == 'secondary'):
            focus = GLOBAL_DATA_POPULATED_SECONDARY
            target_sheet = catalogue_secondary
        for row in range(4, len(focus)+4):
            access = dict()
            for access_col in range(1, len(DATA_CATALOGUE)+1):
                access_column = CATALOGUE_KEYS[access_col-1]
                access_column_data = focus[row-4][access_column]
                access[CATALOGUE_KEYS[access_col-1]] = access_column_data
            for col in range(1, len(DATA_CATALOGUE)):
                column = CATALOGUE_KEYS[col-1]
                column_data = focus[row-4][column]
                if(column == 'lot'):
                    if(column_data != None):
                        if access['invoice_number'] != None:
                            cell = INVOICE_ORDERS[access['invoice_number']]
                        elif column_data in marks_order:
                            cell = column_data
                        else: cell = None
                    else: cell = None
                elif(column == 'invoice_number'):
                    if(column_data != None):
                        cell = column_data.split("||")[0]
                    else:
                        cell = column_data
                elif(column == 'warehouse'):
                    if(column_data != None):
                        cell = column_data
                    else:
                        cell = column_data
                elif(column == 'manufacture_date'):
                    if(column_data != None and column_data != "None"):
                        date = datetime.strptime(
                            column_data,
                            '%Y-%m-%d %H:%M:%S'
                        )
                        # Revert custom date formatting
                        # cellv = DateFormat(date).format("jS M, Y")
                        cellv = date
                        cell = cellv
                    else:
                        cell = column_data
                elif(column == 'company'):
                    # prev = focus[row-4][CATALOGUE_KEYS[col]].replace('.', '')
                    cell = column_data
                    # if(prev != None):
                    #     cell = DatabaseQueryProducerCompany(prev)
                    # else:
                    #     cell = None
                elif(column == 'type'):
                    if(column_data != None):
                        if not re.search('Certified', column_data):
                            val = re.sub("[0-9]+", '', column_data.replace(' ', ''), flags=re.IGNORECASE)
                        else: val = column_data
                        if val == "None":
                            cell = None
                        else: cell = val
                    else:
                        cell = None
                elif(column == 'kgs'):
                    if(column_data == None or column_data[0] == "="):
                        try:
                            # cell = MetaLogic._kgs(access['net'], access['packages']) --> Fallback to bags
                            cell = MetaLogic._kgs(access['net'], access['packages'])
                        except:
                            if(access['packages'] != None and access['net'] != None):
                                cell = MetaLogic._kgs_alt(access['packages'], access['net'])
                            else: cell = None
                    else:
                        cell = column_data
                    focus[row-4][column] = cell
                elif(column == 'tare'):
                    if(column_data == None or column_data[0] == "="):
                        if(access['gross'] != None and access['net'] != None):
                            cell = MetaLogic._tare(access['gross'], access['net'])
                        else: cell = None
                    else:
                        cell = column_data
                    focus[row-4][column] = cell
                elif(column == 'net'):
                    if(column_data == None or column_data[0] == "="):
                        try:
                            cell = MetaLogic._net_alt(access['kgs'], access['bags'])
                        except:
                            if(access['gross'] != None and access['tare'] != None):
                                cell = MetaLogic._net(access['gross'], access['tare'])
                            else: cell = None
                        focus[row-4][column] = cell
                    else:
                        cell = column_data
                else:
                    if(column_data == "None"):
                        cell = None
                    else: cell = column_data
                target_sheet.cell(column=col, row=row, value=cell)
        # Apply post-formatting to rows and cells
        empty_access = ['B','F','M']
        bold=Font(
            size=18,
            bold=True
        )
        for row in range(4, len(focus)+4):
            target_sheet.row_dimensions[row].height = 36.8
        for acc in separator_access:
            for cell in target_sheet[str(acc) + ":" + str(acc)]:
                cell.font = bold
            target_sheet.row_dimensions[acc].height = 42
        for column in empty_access:
            target_sheet.column_dimensions[column].width = 0.75
        target_sheet.active_cell
        for n in range(1, 100):
            target_sheet[('K'+str(n))].number_format = 'dd-mmm'

def GENERATECATALOGUE(input_data, filename, marks_order, reprint_placement):
    global catalogue
    StackGenerator(input_data)
    metadata = CumulativePopulate(LotQuery(), marks_order, reprint_placement)
    
    valuations = ArrangeLots(metadata['catalogue_data'])
    
    PopulateCatalogue.fill('primary', metadata['separator_access']['MAIN_ACCESS'], marks_order)
    PopulateCatalogue.fill('dust', metadata['separator_access']['DUST_ACCESS'], marks_order)
    PopulateCatalogue.fill('secondary', metadata['separator_access']['SECONDARY_ACCESS'], marks_order)
    
    PopulateValuations.fill(valuations)
    
    fs_save_folder = 'media/documents/catalogues/'
    _filename = 'Catalogue_' + filename + '.xlsx'
    dest_save_path = fs_save_folder + _filename
    
    CloseLot(metadata['lotnumber'])
    
    catalogue.save(filename=dest_save_path)
    
    catalogue.close()
    
    return_data = {**metadata, **{'filename': _filename}}
    return return_data
