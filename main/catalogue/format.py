from openpyxl import cell
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM

def readExcel():
    return

font = Font(
    name='Arial Nova Light',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000'
)
fill = PatternFill(
    fill_type=None,
    start_color='FFFFFFFF',
    end_color='FF000000'
)
border = Border(
    left=Side(border_style=None,
                color='FF000000'),
    right=Side(border_style=None,
                color='FF000000'),
    top=Side(border_style=None,
                color='FF000000'),
    bottom=Side(border_style=None,
                color='FF000000'),
    diagonal=Side(border_style=None,
                    color='FF000000'),
    diagonal_direction=0,
    outline=Side(border_style=None,
                    color='FF000000'),
    vertical=Side(border_style=None,
                    color='FF000000'),
    horizontal=Side(border_style=None,
                    color='FF000000')
)
alignment=Alignment(
    horizontal='general',
    vertical='bottom',
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0
)
number_format = 'General'
protection = Protection(
    locked=True,
    hidden=False
)

FORMAT_INITIAL = NamedStyle(
    name="initial",
    font= Font(
        name="Arial"
    ),
    number_format="center"
)

FORMAT_UNSOLD_SALE__AC = NamedStyle(
    name="unsold_sale__ac",
    font = Font(
        name='Arial Nova Light',
        size=11,
        color='FF0000',
        bold=False,
    ),
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    ),
    fill = PatternFill(
        "solid",
        fgColor="FFFFFFFF"
    )
)

FORMAT_CENTER = NamedStyle(
    name="centerhv",
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    )
)

FORMAT_CENTER_GENERAL = NamedStyle(
    name="centerhg",
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    ),
    font=Font(
        size=11,
        bold=False
    ),
)

FORMAT_TITLE = NamedStyle(
    name="title",
    font=Font(
        size=22,
        bold=True
    ),
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    ),
)

FORMAT_SUBTITLE = NamedStyle(
    name="subtitle",
    font=Font(
        size=18,
        bold=True
    ),
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    ),
)

FORMAT_ARIAL_11 = NamedStyle(
    name="arial_nova_light_11",
    font = Font(
        name='Arial Nova Light',
        size=11,
        bold=False,
    ),
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    ),
)

FORMAT_ARIAL_11_WHITE_BG = NamedStyle(
    name="arial_nova_light_11_whitebg",
    font = Font(
        name='Arial Nova Light',
        size=11,
        bold=False,
    ),
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    ),
    fill = PatternFill(
        "solid",
        fgColor="FFFFFFFF"
    )
)

FORMAT_ARIAL_11_BOLD = NamedStyle(
    name="arial_nova_light_11_bold",
    font = Font(
        name='Arial Nova Light',
        size=11,
        bold=True,
    ),
    alignment = Alignment(
        horizontal='center',
        vertical='center'
    ),
)

class Format:
    def fontSize(cell: cell, fontsize: float):
        cell.font += Font(size=fontsize)
    def GeneralCenter(cell: cell):
        cell.style = FORMAT_CENTER_GENERAL
    def formatTitle(cell: cell):
        cell.style = FORMAT_TITLE
    def formatSubTitle(cell: cell):
        cell.style = FORMAT_SUBTITLE
    def formatArial11(cell: cell):
        cell.style = FORMAT_ARIAL_11
    def formatArial11BgWhite(cell: cell):
        cell.style = FORMAT_ARIAL_11_WHITE_BG
    def formatArial11Bold(cell: cell):
        cell.style = FORMAT_ARIAL_11_BOLD
    def formatUnsoldAcSale(cell: cell):
        cell.style = FORMAT_UNSOLD_SALE__AC

medium_border = Border(
    top=Side(border_style=BORDER_MEDIUM, color='00000000'),
    bottom=Side(border_style=BORDER_MEDIUM, color='00000000')
)

subtotals_border = Border(
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000'),
    left=Side(border_style=BORDER_THIN, color='00000000')
)
