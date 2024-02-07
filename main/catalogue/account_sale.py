from openpyxl import load_workbook
from mysql.connector import connect
from main.catalogue.connector import CONNECTOR
import re
from main.catalogue.interpret import TEAGRADES_DATA
from main.catalogue.query import *
from django.core.files.storage import FileSystemStorage
import json
from main.catalogue.format import *
from main.catalogue.helper import *
from datetime import datetime

DATA_LETTERS = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
DATA_ALIAS = {
    "buyer_code": ['buyer'],
    "invoice_number": None,
    "sale_date": None,
    "sale_date_alt": None,
    "prompt_date": None,
    "receiver_address_line1": None,
    "receiver_address_line2": None,
    "receiver_address_line3": None,
    "auction_number": None,
    "lot_number": ['lot no'],
    "invoice_number_buyer": ['invoice'],
    "mark": ['mark'],
    "warehouse": None,
    "packages": ['packages', 'pkgs'],
    "type": ['packing type', 'packaging type'],
    "grade": ['grade'],
    "net": ['net weight', 'net weight(in kg)', 'net kgs'],
    "gross": ['gross weight', 'gross weight(in kg)'],
    "base_price": ['base price', 'price', 'price $'],
    "broker_starting_price": ['broker starting price', 'broker starting price($)'],
    "price": ['sale price', 'sale price($)'],
    "status": ['status'],
    "sold_packages": ['sold packages'],
    "certification": ['certification'],
    "manufacture_date": ['manf date'],
    "producer": ['producer'],
    "broker": ['broker'],
    "number": ["si no", "sl no"],
    "empty1": None,
    "empty2": None,
}

DATA_ACCOUNT_SALE = {
    "buyer_code": 'K5',
    "account_sale_number": 'K11',
    "sale_date": 'E11',
    "sale_date_alt": 'N7',
    "prompt_date": 'H11',
    "receiver_address_line1": 'A5',
    "receiver_address_line2": 'A6',
    "receiver_address_line3": 'A7',
    "auction_number": 'A9',
    "lot_number": 'A15',
    "invoice_number_buyer": 'B15',
    "packages": 'C15',
    "type": 'D15',
    "grade": 'E15',
    "net": 'F15',
    "price": 'G15',
    "empty1": 'I15',
    "empty2": 'K15',
}

DATA_ACCOUNT_SALE_META = {
    "sale_date_alt": 'N7',
    "receiver_address_line1": 'A5',
    "receiver_address_line2": 'A6',
    "receiver_address_line3": 'A7',
    "auction_number": 'A9',
    "auction_number_alt": 'A11',
    "account_sale_number": 'K11',
    "sale_date": 'E11',
    "prompt_date": 'H11',
}

DATA_ACCOUNT_SALE_TOTALS = {
    'pkgs': 'C',
    'net': 'F',
    'amount': 'H',
    'brokerage': 'J',
    'whtax': 'L',
    'payable_amount': 'M',
}

DATA_ACCOUNT_SALE_TAX_SUMMARY = {
    'amount': 'A',
    'brokerage': 'B',
    'gross': 'C',
    'whtax': 'E',
    'payable_amount': 'F',
}

DATA_SALE_RELATION = {
    'lot_number': 'A',
    'invoice_number_buyer': 'B',
    'packages': 'C',
    'type': 'D',
    'grade': 'E',
    'net': 'F',
    'price': 'G',
    '_amount': 'H',
    'empty1': 'I',
    '_brokerage': 'J',
    'empty2': 'K',
    '_whtax': 'L',
    '_payable_amount': 'M',
}

DATA_SALE = {
    'lot_number': 'A',
    'invoice_number_buyer': 'B',
    'packages': 'C',
    'type': 'D',
    'grade': 'E',
    'net': 'F',
    'price': 'G',
    '_amount': '=F*G',
    'empty1': 'I',
    '_brokerage': '=H*-0.75%',
    'empty2': 'K',
    '_whtax': '=J*-5%',
    '_payable_amount': '=H+J+L',
}

# Merge M and N

def AccountSaleNumberQuery():
    try:
        with connect(**CONNECTOR) as connection:
            current_number = '''SELECT `number` FROM `account_sale_number`'''
            with connection.cursor() as cursor:
                cursor.execute(current_number)
                row = cursor.fetchone()
                return row[0]

    except Error as e:
            print(e)

ACCOUNT_SALE_COUNTER = AccountSaleNumberQuery()
def AcSaleCounter():
    global ACCOUNT_SALE_COUNTER
    if(ACCOUNT_SALE_COUNTER < 2000):
        ACCOUNT_SALE_COUNTER += 1
        if len(str(ACCOUNT_SALE_COUNTER)) == 1:
            ACCOUNT_SALE_COUNTER = '00' + str(ACCOUNT_SALE_COUNTER)
        elif len(str(ACCOUNT_SALE_COUNTER)) == 2:
            ACCOUNT_SALE_COUNTER = '0' + str(ACCOUNT_SALE_COUNTER)
    else:
        ACCOUNT_SALE_COUNTER = 1
        ACCOUNT_SALE_COUNTER = '00' + str(1)
    return ACCOUNT_SALE_COUNTER

def CloseAccountSaleNumber(account_sale_number, Pid):
    try:
        with connect(**CONNECTOR) as connection:
            update_lot = '''UPDATE `main_auctions` SET `account_sale_number` = %s WHERE `Pid` = %s'''
            with connection.cursor() as cursor:
                cursor.execute(update_lot, (account_sale_number, Pid,))
                connection.commit()

    except Error as e:
        print(e)

def GenerateAccountSaleNumber(auction_number, auction_number_alt, auction_year, _auction_number, counter):
    base = 'PRME/'
    base_conform = 'PRME_'
    return {
        'number': base + str(auction_number) + '/' + Helper.number_prefix(counter),
        'number_alt': base + str(auction_number_alt)  + '/' + Helper.number_prefix(counter),
        'file': base_conform + str(auction_number) + '_' + Helper.number_prefix(counter),
        'ac_sale': base + str(last2(auction_year)) + '/' + str(_auction_number)
    }
    # return {
    #     'number': base + str(auction_number) + '/' + str(AcSaleCounter()),
    #     'file': base_conform + str(auction_number) + '_' + str(AcSaleCounter())
    # }

def findAlias(list, value):
    perfect = [
        'grade',
        'mark',
        "number",
    ]
    end = len(list)
    counter = 1
    for alias in list:
        if(value in perfect):
            if(re.search(alias, value) != None):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        else:
            if(value == alias):
                return True
            else:
                if(counter == end):
                    return False
                else:
                    pass
        counter += 1
    

def getAliasRelation(value):
    endif = 0
    for alias in DATA_ALIAS:
        endif += 1
        if(DATA_ALIAS[alias] != None):
            if(findAlias(DATA_ALIAS[alias], value.lower())):
                return alias
            else:
                if endif == len(DATA_ALIAS):
                    return alias
                else:
                    continue
        else: continue

def last2(s):
    s = str(s)
    length = len(s)
    val = ''
    counter = 0
    for v in s:
        if(counter >= length-2):
            val += s[counter]
        counter += 1
    return val
class DataInterpretor:
    def init_data(left_bound, data_layer, file):
        left_bound = int(left_bound)
        data_layer = int(data_layer)
        folder='media/documents/sales/'
        fs = FileSystemStorage(location=folder)
        file_data = fs.open(file, 'rb+')
        WORKBOOK = load_workbook(filename = file_data )
        DATA = {}
        RELATION = []
        sheet = WORKBOOK.active
        bc = 0
        inner = list()
        endif_check = list()
        for row in sheet.values:
            allow = True
            inner_counter = 0
            for value in row:
                if(inner_counter >= left_bound-1):
                    if(value != None):
                        if(bc == data_layer-1):
                            value = re.sub(r'\t', '', value)
                            RELATION.append(value)
                        inner.append(value)
                    else:
                        inner.append(None)
                        endif_check.append(value)
                inner_counter += 1
            if(len(endif_check) >= 5 and bc >= 5):
                allow = False
                break
            elif(bc >= data_layer and len(endif_check) < 5):
                if allow: DATA[bc-(data_layer)] = row[slice(left_bound-1, len(row))]
            inner = list()
            endif_check = list()
            bc += 1
        return {
            'relation': RELATION,
            'data': DATA
        }

    def generate_data(data):
        set_data = list()
        counter = 0
        for vals in data['data'].values():
            inner_counter = 0
            inner_dict = dict()
            for inner_data in vals:
                relation = getAliasRelation(data['relation'][inner_counter])
                if(relation != False):
                    inner_dict[relation] = inner_data
                else:
                    inner_dict[relation] = None
                inner_counter += 1
            set_data.append(inner_dict)
            counter += 1
        return set_data

STACK_DATA = {}
combined = list()
LOT_STATUS_RELATION = dict()
LOT_MARK_RELATION = dict()
populated = list()
PRODUCERS = list()
PRODUCERS_RELATION = dict()
PRODUCERS_OUTLOT = list()

def FormatOutlot(lot):
    retain = ['lot_number', 'invoice_number_buyer', 'sold_packages', 'pkgs', 'packages', 'type', 'grade', 'net', 'price']
    if lot:
        lot['price'] = 'UNSOLD'
        for value in lot:
            if value not in retain:
                lot[value] = None
    return lot

def StackGenerator(input_data, catalogue_data):
    global STACK_DATA
    global PRODUCERS
    global PRODUCERS_RELATION
    global PRODUCERS_OUTLOT
    global LOT_MARK_RELATION
    combined = list()
    counter = 0
    for file in input_data:
        STACK_DATA[counter] = DataInterpretor.generate_data(
            DataInterpretor.init_data(
                file['left_bound'],
                file['data_layer'],
                file['file']
            )
        )
        counter += 1
    for data in STACK_DATA.values():
        combined += data
    STACK_DATA = combined
    for lot in STACK_DATA:
        LOT_STATUS_RELATION[lot['lot_number']] = lot['status']
        LOT_MARK_RELATION[lot['lot_number']] = {
            'mark': lot['mark'],
            'invoice': lot['invoice_number_buyer'],
        }
    for data in STACK_DATA:
        exist = list()
        inner_val = dict()
        for single in data:
            if(single in DATA_ACCOUNT_SALE.keys()):
                exist.append(single)
        for value in DATA_ACCOUNT_SALE.keys():
            if(value in exist):
                inner_val[value] = data[value]
            else:
                inner_val[value] = None
        populated.append(inner_val)
    
    def StackAccess(l, v, mark):
        lval = list()
        for lot in STACK_DATA:
            for val in lot:
                if val == "invoice_number_buyer":
                    # if str(f'{lot[val]}||{mark}') == l:
                    if replaceResale(lot[val]) == l:
                        lval = lot
                        return lval[v]
    
    folder='media/documents/catalogue_data'
    fsc = FileSystemStorage(location=folder)
    
    with fsc.open(catalogue_data, 'rb+') as fcc_file:
        file_datac = json.load(fcc_file)
        
    for lot in populated:
        mark = LOT_MARK_RELATION[lot['lot_number']]['mark']
        lot['producer'] = GetAcSaleProducerMark(file_datac, replaceResale(lot['invoice_number_buyer']), mark)
        if replaceResale(lot['invoice_number_buyer']) != '':
            mark_from_sale = GetAcSaleProducerMark(file_datac, replaceResale(lot['invoice_number_buyer']), mark)
            if mark_from_sale != None:
                PRODUCERS.append(re.sub(r'[0-9]', '', GetAcSaleProducerMark(file_datac, replaceResale(lot['invoice_number_buyer']), mark)))
            else:
                PRODUCERS.append(re.sub(r'[0-9]', '', StackAccess(replaceResale(lot['invoice_number_buyer']), "mark", mark)))

    PRODUCERS = list(set(PRODUCERS))
    print(PRODUCERS)
    for producer in PRODUCERS:
        PRODUCERS_RELATION[producer] = list()
    for producer in PRODUCERS:
        for lot in populated:
            if lot['producer'] == producer:
                if LOT_STATUS_RELATION[lot['lot_number']].lower() == 'sold':
                    PRODUCERS_RELATION[producer].append(lot)
                else:
                    PRODUCERS_RELATION[producer].append(FormatOutlot(lot))

    for lot in populated:
        if LOT_STATUS_RELATION[lot['lot_number']].lower() == 'outlot':
            number = lot['lot_number']
            for _lot in STACK_DATA:
                if(_lot['lot_number']) == number:
                    PRODUCERS_OUTLOT.append(_lot)
    
# def arrangeData():
#     global BUYERS
#     for lot in populated:
#         if lot['buyer_code'] != '': BUYERS.append(lot['buyer_code'])
#     BUYERS = set(BUYERS)
#     for buyer in BUYERS:
#         BUYERS_RELATION[buyer] = list()
#     for buyer in BUYERS:
#         for lot in populated:
#             if lot['buyer_code'] == buyer and LOT_STATUS_RELATION[lot['lot_number']].lower() == 'sold':
#                 BUYERS_RELATION[buyer].append(lot)
#             else:
#                 BUYERS_OUTLOT.append(lot)
# arrangeData()

def formatAddress(address):
    def splitAddress(val):
        data = re.split("[,]+", str(val), flags=re.IGNORECASE)
        if(data[1][0] == ' '):
            data[1] = re.sub(' ', '', data[1], 1)
        return data
    if(address !=  None):
        if(re.search(',', address)):
            # P.O. Box search
            if(re.search('P.O.Box', address)):
                address = address.replace('P.O.Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address = address.replace('P.O. Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address == address
            elif(re.search('Box', address)):
                address = address.replace('Box', 'P.O. BOX')
            else:
                address = 'P.O. BOX ' + address
            # (-) Hyphen search
            if(re.search(' - ', address)):
                address == address
            elif(re.search('- ', address)):
                address = address.replace('- ', ' - ')
            elif(re.search(' -', address)):
                address = address.replace(' -', ' - ')
            elif(re.search('-', address)):
                address = address.replace('-', ' - ')
            # Format
            address = address.upper()
            return splitAddress(address)
        else:
            if(re.search('Mombasa', address)):
                address = address.replace('Mombasa', ', MOMBASA')
            elif(re.search('Nairobi', address)):
                address = address.replace('Nairobi', ', NAIROBI')
            elif(re.search('Kericho', address)):
                address = address.replace('Kericho', ', KERICHO')
            else:
                address += ', NAIROBI'
            # P.O. Box search
            if(re.search('P.O.Box', address)):
                address = address.replace('P.O.Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address = address.replace('P.O. Box', 'P.O. BOX')
            elif(re.search('P.O. Box', address)):
                address == address
            elif(re.search('Box', address)):
                address = address.replace('Box', 'P.O. BOX')
            else:
                address = 'P.O. BOX ' + address
            # (-) Hyphen search
            if(re.search(' - ', address)):
                address == address
            elif(re.search('- ', address)):
                address = address.replace('- ', ' - ')
            elif(re.search(' -', address)):
                address = address.replace(' -', ' - ')
            elif(re.search('-', address)):
                address = address.replace('-', ' - ')
            # Format
            address = address.upper()
            return splitAddress(address)
    else:
        return ['', '']
    
def populate_number(val, level):
    data_length = len(val)
    functions = [
        'SUM', 'PRODUCT', 'DIFFERENCE', 'AVG',
    ]
    hasfn = False
    for fn in functions:
        if(re.search(fn, val)):
            hasfn = True
    counter = 0
    brack = False
    for v in val:
        if(v == '('):
            brack = True
        if v in DATA_LETTERS:
            if(not hasfn):
                val = val.replace(v, v+str(level))
            else:
                if(brack == True):
                    val = val.replace(v, v+str(level))
        if(counter == data_length-1):
            return val
        counter += 1

def replaceResale(v):
    resalestr = ["Resale", "resale", "RESALE"]
    for resale in resalestr:
        if resale in str(v):
            head, sep, tail = str(v).partition(resale)
            head += sep
            return str(head).replace(" ", "").replace("-Resale", "").replace("-resale", "").replace("-RESALE", "").replace("Resale", "").replace("resale", "").replace("RESALE", "")
        else:
            return v

NUMBER_FORMAT_CELLS = list()
def PopulateRow(sheet, level, row_data, catalogue_data):
    global NUMBER_FORMAT_CELLS
    NUMBER_FORMAT_CELLS = list()
    for data in DATA_SALE:
        mark = LOT_MARK_RELATION[row_data['lot_number']]["mark"]
        # warehouse = DatabaseQueryProducerCompany(mark)
        # mark = None
        if(data[0] != '_'):
            # if(data == 'warehouse'):
            #     sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = warehouse
            if(data == 'warehouse'):
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = GetAcSaleProducerMark(catalogue_data, replaceResale(row_data['invoice_number_buyer']), mark)
            elif(data == 'packages' or data == 'net'):
                if row_data[data] != None:
                    sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = int(row_data[data])
            elif(data == 'invoice_number_buyer'):
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = replaceResale(row_data[data])
            else:
                sheet[str(str(DATA_SALE_RELATION[data])+str(level))] = row_data[data]
            Format.formatArial11BgWhite(sheet[str(str(DATA_SALE_RELATION[data])+str(level))])
        else:
            # CHECK Outlot Status
            if LOT_STATUS_RELATION[row_data['lot_number']] == 'Sold':
                sheet[populate_number(DATA_SALE_RELATION[data], level)] = populate_number(DATA_SALE[data], level)
                sheet[populate_number(DATA_SALE_RELATION[data], level)].number_format = '$#,##0.00'
                if data == '_payable_amount':
                    sheet.merge_cells(DATA_SALE_RELATION[data] + str(level) + ':N' + str(level))
                Format.formatArial11BgWhite(sheet[populate_number(DATA_SALE_RELATION[data], level)])
                cell = populate_number(DATA_SALE_RELATION[data], level)
                NUMBER_FORMAT_CELLS.append(cell)
            else:
                targets = ['A','B','C','D','E','F','G']
                for target in targets:
                    Format.formatUnsoldAcSale(sheet[target+str(level)])

    return NUMBER_FORMAT_CELLS

def GetAcSaleProducerMark(catalogue_data, invoice_number, mark):
    lot_producer = None
    for data in catalogue_data:
        for lot in data:
            if(lot == 'invoice_number'):
                if(str(data["invoice_number"]) == str(f'{invoice_number}||{mark}')):
                    lot_producer = data['mark']
    return lot_producer

def GenerateAccountSale(data, custom_values, counter, producer):
    folder='media/resources/'
    fs = FileSystemStorage(location=folder)
    
    catalogue_data = custom_values['catalogue_data']

    folder='media/documents/catalogue_data'
    fsc = FileSystemStorage(location=folder)
    
    with fsc.open(catalogue_data, 'rb+') as fcc_file:
        file_datac = json.load(fcc_file)

    LOT_COMBINED = {producer: list()}

    for producer in data:
        lot = data[producer]
        for lot_data in lot:
            LOT_COMBINED[producer].append(lot_data)

    def _Generate(data, producer, custom_values, type="default"):
        folder='media/resources/'
        fs = FileSystemStorage(location=folder)
        template_default = fs.open('ACCOUNT SALE TEMPLATE.xlsx', 'rb+')

        catalogue_data = custom_values['catalogue_data']

        folder='media/documents/catalogue_data'
        alt_fsc = FileSystemStorage(location=folder)
        
        with alt_fsc.open(catalogue_data, 'rb+') as _fcc_file:
            alt_file_datac = json.load(_fcc_file)

        file_data = template_default
        ACsale_template = load_workbook(filename = file_data)

        ac_sale_file = GenerateAccountSaleNumber(custom_values['auction_number'], custom_values['auction_number_alt'], custom_values['auction_year'], custom_values['_auction_number'], counter)['file']

        account_sale_values = GenerateAccountSaleNumber(custom_values['auction_number'], custom_values['auction_number_alt'], custom_values['auction_year'], custom_values['_auction_number'], counter)
        ac_sale_data = custom_values['account_sale_data']
        account_sale_number_alt = account_sale_values['number_alt']

        ac_sale_number = account_sale_values['ac_sale']

        sale_date = custom_values['sale_date']
        prompt_date = custom_values['prompt_date']

        fs_save_folder = 'media/documents/account_sales/'
        _filename = 'AccountSale_' + '(' + producer + ')' + ac_sale_file + '.xlsx'
        dest_save_path = fs_save_folder + _filename

        for producer in data:
            producer_search = producer
            producer_company = DatabaseQueryProducerCompany(producer)
            producer_address = DatabaseQueryProducerAddress(producer_search)
            code = producer
            address_line1 = producer_company.upper()
            address_line2 = formatAddress(producer_address)[0]
            address_line3 = formatAddress(producer_address)[1]
            # Auction Date Title
            auction_date_start = "AUCTION No. " + str(custom_values['auction_number_alt']) + " of "
            auction_date_end = ", " + custom_values['auction_year']
            meta_relation = {
                'account_sale_number': ac_sale_number,
                'sale_date':  sale_date,
                'sale_date_alt': sale_date,
                'prompt_date': prompt_date,
                'receiver_address_line1': address_line1,
                'receiver_address_line2': address_line2,
                'receiver_address_line3': address_line3,
                'auction_number': auction_date_start + custom_values['auction_date_title'] + auction_date_end,
                'auction_number_alt': custom_values['auction_number_alt']
            }
            lot_main = list()
            lot_secondary = list()
            lot_start = 15
            lot_initial = lot_start
            lot_secondary_title = 16
            lot_secondary_start = 17
            totals = 19
            tax_summary = 23
            lot = data[producer]

            for sale in lot:
                for value in sale:
                    if value == 'grade':
                        if TEAGRADES_DATA[sale[value]] == 'primary':
                            lot_main.append(sale)
                        else:
                            lot_secondary.append(sale)

            # Extract Catalogue Lot Order
            LOT_ORDER = list()
            for data in alt_file_datac:
                for lot in data:
                    if(lot == 'lot'):
                        LOT_ORDER.append(data["lot"])
            
            lot_main = sorted(lot_main, key = lambda x: LOT_ORDER.index(int(x['lot_number'])))
            lot_secondary = sorted(lot_secondary, key = lambda y: LOT_ORDER.index(int(y['lot_number'])))

            main_length = len(lot_main)
            secondary_length = len(lot_secondary)
            data_length = (main_length)+(secondary_length)
            if main_length != 0:
                lot_secondary_ac_start = lot_secondary_start+(main_length-1)
            else:
                lot_secondary_ac_start = lot_secondary_start
            lot_secondary_title = lot_secondary_title+(main_length-1)
            lot_end = lot_secondary_ac_start
            if main_length != 0:
                totals += (main_length-1)
                tax_summary += (main_length-1)
            if secondary_length != 0:
                totals += (secondary_length-1)
                tax_summary += (secondary_length-1)
                lot_end += (secondary_length-1)
            balance_proceeds = tax_summary+1
            total_charges = tax_summary
            other_charges = tax_summary-1
            warehouse_charges = tax_summary-2
            crates_charges = tax_summary-3

            # DEBUG VALUES
            # ("PROMPT DATE => " + str(prompt_date))
            # ("LOT SECONDARY START => " + str(lot_secondary_ac_start))
            # ("LOT SECONDARY TITLE => " + str(lot_secondary_title))
            # ("TOTALS => " + str(totals))
            # ("SUMMARY => " + str(tax_summary))
            # ("MAIN LENGTH => " + str(main_length))
            # ("SECONDARY LENGTH => " + str(secondary_length))
            # ("DATA LENGTH => " + str(data_length))

            if(data_length > 1):
                if main_length > 1:
                    ACsale_template.active.insert_rows(lot_start, main_length-1)
                if secondary_length > 1:
                    ACsale_template.active.insert_rows(lot_secondary_ac_start, secondary_length-1)
            NUMBER_CELLS = list()
            for lot_data in lot_main:
                NUMBER_CELLS = [*NUMBER_CELLS, *PopulateRow(ACsale_template.active, lot_start, lot_data, file_datac)]
                lot_start += 1
            for lot_data in lot_secondary:
                NUMBER_CELLS = [*NUMBER_CELLS, *PopulateRow(ACsale_template.active, lot_secondary_ac_start, lot_data, file_datac)]
                lot_secondary_ac_start += 1
            for cell in NUMBER_CELLS:
                if re.search('J', cell):
                    ACsale_template.active[cell].number_format = '0.00_);(0.00)'
                else:
                    ACsale_template.active[cell].number_format = '#,##0.00'
            SUMMARY_RELATION = {}

            for total in DATA_ACCOUNT_SALE_TOTALS:
                ACsale_template.active[DATA_ACCOUNT_SALE_TOTALS[total]+str(totals)] = '=SUM(' + DATA_ACCOUNT_SALE_TOTALS[total] + str(lot_initial) + ':' + DATA_ACCOUNT_SALE_TOTALS[total] + str(lot_end) + ')'
                SUMMARY_RELATION[total] = '=' + DATA_ACCOUNT_SALE_TOTALS[total]+str(totals)

            for summary in DATA_ACCOUNT_SALE_TAX_SUMMARY:
                if(summary != 'gross' and summary != 'payable_amount'):
                    ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)] = SUMMARY_RELATION[summary]
                    if(summary == 'amount' or summary == 'whtax'):
                        ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)].number_format = '#,##0.00'
                elif(summary == 'gross'):
                    ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)] = '=SUM(' + DATA_ACCOUNT_SALE_TAX_SUMMARY['amount'] + str(tax_summary) + ':' + DATA_ACCOUNT_SALE_TAX_SUMMARY['brokerage'] + str(tax_summary) + ')'
                    ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)].number_format = '0.00_);(0.00)'
                elif(summary == 'payable_amount'):
                    ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)] = '=SUM(' + DATA_ACCOUNT_SALE_TAX_SUMMARY['gross'] + str(tax_summary) + ':' + DATA_ACCOUNT_SALE_TAX_SUMMARY['whtax'] + str(tax_summary) + ')'
                    ACsale_template.active[DATA_ACCOUNT_SALE_TAX_SUMMARY[summary]+str(tax_summary)].number_format = '#,##0.00'
            for meta in DATA_ACCOUNT_SALE_META:
                if meta == 'prompt_date':
                    date = datetime.strptime(
                        meta_relation[meta],
                        '%d/%m/%Y'
                    )
                    ACsale_template.active[DATA_ACCOUNT_SALE_META[meta]] = date
                else:
                    ACsale_template.active[DATA_ACCOUNT_SALE_META[meta]] = meta_relation[meta]

            # BALANCE PROCEEDS PLACEMENT
            ACsale_template.active['M'+str(total_charges)] = '=SUM(M' + str(crates_charges) + ':' + 'M' + str(other_charges) + ')'
            ACsale_template.active['M'+str(balance_proceeds)] = '=ABS(M' + str(total_charges) + '-' + 'M' + str(totals) + ')'
            ACsale_template.active['M'+str(balance_proceeds)].number_format = '#,##0.00'

            ##### POST-FORMATTING
            ACsale_template.active.merge_cells('C' + str(tax_summary) + ':D' + str(tax_summary))
            ACsale_template.active.merge_cells('M' + str(tax_summary) + ':N' + str(tax_summary))
            ACsale_template.active.merge_cells('M' + str(balance_proceeds) + ':N' + str(balance_proceeds))
            ACsale_template.active.merge_cells('M' + str(totals) + ':N' + str(totals))
            ACsale_template.active[DATA_ACCOUNT_SALE_META["prompt_date"]].number_format = 'dd-mmm-yy'
            
            ACsale_template.active.title = ac_sale_file
            
            ACsale_template.save(filename=dest_save_path)
            
            return _filename

    _dir = None
    if len(LOT_COMBINED[producer]) >= 1:
        _dir = _Generate(LOT_COMBINED, producer, custom_values, "default")
    
    return {
        'file': _dir,
        'counter': counter,
    }

class PopulateAccountSale():
    def fill_lots(custom_values):
        counter = int(custom_values['account_sale_number'])
        dirs = list()
        for producer in PRODUCERS_RELATION:
            producer_dirs = GenerateAccountSale({
                producer: PRODUCERS_RELATION[producer]
            }, custom_values, counter, producer)
            print(producer_dirs)
            if producer_dirs['file']:
                dirs.append(producer_dirs['file'])
            counter = producer_dirs['counter']
            counter += 1
        CloseAccountSaleNumber(counter, custom_values['auction_Pid'])
        return {
            'dirs': dirs
        }

def ACCOUNTSALEGENERATOR(input_data, custom_data):
    StackGenerator(input_data, custom_data['catalogue_data'])
    return PopulateAccountSale.fill_lots(custom_data)
